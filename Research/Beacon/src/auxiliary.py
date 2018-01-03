#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct 31 18:03:35 2017

auxiliary function for paper

@author: eleme-yi
"""

import os 
import openpyxl
import numpy as np
from openpyxl import Workbook
from os import listdir
from os.path import isfile, join

def gather_rr_edge_files():
    dir_path = os.path.dirname(os.path.realpath(__file__))
    my_path = os.path.join(dir_path, 'data/time_map/')
    # Joint several individual excel files for training
    
    only_files = [f for f in listdir(my_path) if isfile(join(my_path, f))]
    
    output_file = os.path.join(my_path,'rr_edge_all.xlsx')
    
    wb_all = Workbook()
    ws1 = wb_all.active
    ws1.title = "Sheet0"
    
    current_row = 0
    is_first_file = True
    num_file_read = 0
    
    for file_name in only_files:
        file_name_list = file_name.split('_')
        if len(file_name_list) < 3:
            continue
        date_str = file_name_list[2]
        date_str_list = date_str.split('-')
        if file_name_list[0] == 'rr' and file_name_list[1] == 'edge' \
        and date_str_list[0] == '2017':
            num_file_read = num_file_read + 1
            print(num_file_read, 'file read!')
            # Read data and write to 'all' file
            file_path = os.path.join(my_path,file_name)        
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.get_sheet_by_name('Sheet0')
            row_count = sheet.max_row
            column_count = sheet.max_column
    
            if is_first_file:
                first_row = 1
                is_first_file = False
            else:
                first_row = 2
            for i in range(first_row,row_count+1):
                current_row = current_row + 1
                for j in range(1,column_count+1):
                    element = sheet.cell(row=i,column=j).value
                    ws1.cell(row = current_row, column = j).value = element
                        
    wb_all.save(filename = output_file)
    print(num_file_read,'files read in total.')
    
def read_rr_edge_file(date):
    if date == 'all':
        rr_edge_file = 'data/time_map/rr_edge_all'
    else:
        rr_edge_file = '_'.join(['data/time_map/rr_edge',str(date)])    
        # rr_edge_file = '_'.join(['data/time_map/rr_edge',str(rider_id)])
    
    wb = openpyxl.load_workbook(rr_edge_file+'.xlsx')
    sheet = wb.get_sheet_by_name('Sheet0')
    
    row_count = sheet.max_row
    column_count = sheet.max_column
    
    data_num = row_count - 1
    feature_num = 9
    
    feature_name = [sheet.cell(row=1,column=i+1).value for i in range(0,column_count)]
    from_lat_index = feature_name.index('from_lat')
    from_lon_index = feature_name.index('from_lon')
    from_floor_index = feature_name.index('from_floor')
    to_lat_index = feature_name.index('to_lat')
    to_lon_index = feature_name.index('to_lon')
    to_floor_index = feature_name.index('to_floor')
    distance_delta_index = feature_name.index('distance_delta')
    hour_index = feature_name.index('hour')
    weather_index = feature_name.index('weather')
    time_delta_index = feature_name.index('time_delta')
    
    X = np.empty([data_num,feature_num])
    y = np.empty([data_num,1])
    for i in range(0,data_num):
        X[i][0] = sheet.cell(row=i+2,column=from_lat_index+1).value             # lat1
        X[i][1] = sheet.cell(row=i+2,column=from_lon_index+1).value             # lon1
        X[i][2] = sheet.cell(row=i+2,column=from_floor_index+1).value           # floor1
        X[i][3] = sheet.cell(row=i+2,column=to_lat_index+1).value               # lat2
        X[i][4] = sheet.cell(row=i+2,column=to_lon_index+1).value               # lon2
        X[i][5] = sheet.cell(row=i+2,column=to_floor_index+1).value             # floor2
        X[i][6] = sheet.cell(row=i+2,column=distance_delta_index+1).value*1000  # distance
        X[i][7] = sheet.cell(row=i+2,column=hour_index+1).value                 # hour
        X[i][8] = sheet.cell(row=i+2,column=weather_index+1).value              # weather
        y[i] = sheet.cell(row=i+2,column=time_delta_index+1).value              # duration
    
    return X,y

def read_rc_edge_file(date):
    if date == 'all':
        rr_edge_file = 'data/time_map/rc_edge_all'
    else:
        rr_edge_file = '_'.join(['data/time_map/rc_edge',str(date)])    
        # rr_edge_file = '_'.join(['data/time_map/rr_edge',str(rider_id)])
    
    wb = openpyxl.load_workbook(rr_edge_file+'.xlsx')
    sheet = wb.get_sheet_by_name('Sheet0')
    
    row_count = sheet.max_row
    column_count = sheet.max_column
    
    data_num = row_count - 1
    feature_num = 4
    
    feature_name = [sheet.cell(row=1,column=i+1).value for i in range(0,column_count)]
    #from_lat_index = feature_name.index('from_lat')
    #from_lon_index = feature_name.index('from_lon')
    #from_floor_index = feature_name.index('from_floor')
    #to_lat_index = feature_name.index('to_lat')
    #to_lon_index = feature_name.index('to_lon')
    #to_floor_index = feature_name.index('to_floor')
    distance_delta_index = feature_name.index('order_distance')
    #hour_index = feature_name.index('hour')
    weather_index = feature_name.index('weather_code')
    grid_avg_lead_time_index = feature_name.index('grid_avg_lead_time_30')
    rst_avg_lead_time_index = feature_name.index('rst_avg_lead_time_30')
    time_delta_index = feature_name.index('label')
    
    X = np.empty([data_num,feature_num])
    y = np.empty([data_num,1])
    for i in range(0,data_num):
        X[i][0] = sheet.cell(row=i+2,column=distance_delta_index+1).value*1000  # distance
        X[i][1] = sheet.cell(row=i+2,column=weather_index+1).value              # weather
        X[i][2] = sheet.cell(row=i+2,column=grid_avg_lead_time_index+1).value   # grid average
        X[i][3] = sheet.cell(row=i+2,column=rst_avg_lead_time_index+1).value    # restaurant average
        y[i] = sheet.cell(row=i+2,column=time_delta_index+1).value              # duration
    
    return X,y