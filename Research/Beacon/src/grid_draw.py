#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 13 15:58:30 2017

@author: eleme-yi
"""

# Draw grid

import os 
import matplotlib.pyplot as plt
import gmplot
import openpyxl

dir_path = os.path.dirname(os.path.realpath(__file__))
my_path = os.path.join(dir_path, 'data/grid/')

sep = '_'

# Draw shop
#file_name = "indoor_shop_gps.xlsx"
#output_file = sep.join(["figures/indoor_shop",".html"])
#lat_col = 3
#lon_col = 4

#file_name = "cluster_gps.xlsx"
#output_file = sep.join(["figures/all_cluster_shop",".html"]) 
#lat_col = 2
#lon_col = 3
#
#file_path = os.path.join(my_path,file_name)        
#wb = openpyxl.load_workbook(file_path)
#sheet = wb.get_sheet_by_name('Sheet0')
#
#indoor_shop_latitude = []
#indoor_shop_longitude = []
#
#row_count = sheet.max_row
#column_count = sheet.max_column
#
#for index in range(2,int(row_count/2)):
#    i = index * 2
#    lat = sheet.cell(row=i,column=lat_col).value
#    lon = sheet.cell(row=i,column=lon_col).value
#    indoor_shop_latitude.append(lat)
#    indoor_shop_longitude.append(lon)
#
#gmap = gmplot.GoogleMapPlotter(31.2325, 121.381895, 16)
##gmap = gmplot.from_geocode("Shanghai")
##gmap.plot(indoor_shop_latitude, indoor_shop_longitude, 'blue', edge_width=10)
#gmap.scatter(indoor_shop_latitude, indoor_shop_longitude, 'blue', marker=True)
#gmap.draw(output_file)

# Draw Grid
central_file_name = "central_grid_polygon.xlsx"
noncentral_file_name = "noncentral_grid_polygon.xlsx"
output_file = sep.join(["figures/grid_chosen",".html"]) 

central_file_path = os.path.join(my_path,central_file_name)
noncentral_file_path = os.path.join(my_path,noncentral_file_name)
central_wb = openpyxl.load_workbook(central_file_path)
noncentral_wb = openpyxl.load_workbook(noncentral_file_path)
central_sheet = central_wb.get_sheet_by_name('Sheet0')
noncentral_sheet = noncentral_wb.get_sheet_by_name('Sheet0')




file_name_list = ["inner_grid_polygon.xlsx","middle_grid_polygon.xlsx","outter_grid_polygon.xlsx"]
color_list = ['blue','cyan','red']

gmap = gmplot.GoogleMapPlotter(31.2325, 121.381895, 16)

for index in range(0,len(file_name_list)):
    
    file_name = file_name_list[index]
    color = color_list[index]
    
    file_path = os.path.join(my_path,file_name)
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name('Sheet0')
    
    row_count = central_sheet.max_row
    
    for i in range(2, row_count):
        lat = []
        lon = []    
        poly_str = sheet.cell(row=i,column=3).value
        poly_list = poly_str.split(';')
        cent_str = sheet.cell(row=i,column=2).value
        cent_pos_list = cent_str.split(',')
        cent_lat = float(cent_pos_list[0])
        cent_lon = float(cent_pos_list[1])
        gmap.scatter([cent_lat], [cent_lon], color, marker=True)
        for j in range(0,len(poly_list)):
            pos_list = poly_list[j].split(',')
            lat.append(float(pos_list[0]))
            lon.append(float(pos_list[1]))
            gmap.plot(lat, lon, color, edge_width=10)
    
# Some missing grids        
gmap.scatter([31.2246243], [121.4839845], 'red', marker=True)        
    
gmap.draw(output_file)