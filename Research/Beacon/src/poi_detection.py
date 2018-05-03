'''
POI inaccuracy detection from Phase III data
Yi DING
3/10/18

'''

import os
import openpyxl
import gmplot
import numpy as np

## Draw shops whose beacon are heard at the same second
# This path of current file
dir_path = os.path.dirname(os.path.realpath(__file__))

data_path = '/Users/eleme-yi/Downloads/beacon-temp'
figure_path = '/Users/eleme-yi/Downloads/beacon-temp/suspect_shop_figures'
file_name = 'poi_detection_by_rider_cnt.xlsx'
file_path = os.path.join(data_path, file_name)

# Open workbook
wb = openpyxl.load_workbook(file_path)
sheet = wb.get_sheet_by_name('Sheet0')
row_count = sheet.max_row
column_count = sheet.max_column

# Get argument list
argument_list = []
for j in range(1, column_count+1):
    argument_list.append(sheet.cell(row=1,column=j).value)
print(argument_list)

# Get indices
index_distance = argument_list.index('distance')
index_shop_1 = argument_list.index('shop_id_1')
index_shop_2 = argument_list.index('shop_id_2')
index_lat_1 = argument_list.index('shop_latitude_1')
index_lon_1 = argument_list.index('shop_longitude_1')
index_lat_2 = argument_list.index('shop_latitude_2')
index_lon_2 = argument_list.index('shop_longitude_2')
index_rider_cnt = argument_list.index('rider_cnt')

# Read in data
# Get lists from raw file
list_shop_1 = []
list_lat_1 = []
list_lon_1 = []
list_shop_2 = []
list_lat_2 = []
list_lon_2 = []
list_distance = []
list_rider_cnt = []
list_all_shops = []
for i in range(2, row_count+1):
    shop_1_id = sheet.cell(row=i, column=index_shop_1 + 1).value
    shop_2_id = sheet.cell(row=i, column=index_shop_2 + 1).value
    list_shop_1.append(shop_1_id)
    list_shop_2.append(shop_2_id)
    list_lat_1.append(sheet.cell(row=i, column=index_lat_1 + 1).value)
    list_lon_1.append(sheet.cell(row=i, column=index_lon_1 + 1).value)
    list_lat_2.append(sheet.cell(row=i, column=index_lat_2 + 1).value)
    list_lon_2.append(sheet.cell(row=i, column=index_lon_2 + 1).value)
    list_distance.append(sheet.cell(row=i, column=index_distance + 1).value)
    list_rider_cnt.append(sheet.cell(row=i, column=index_rider_cnt + 1).value)
    if shop_1_id not in list_all_shops:
        list_all_shops.append(shop_1_id)
    if shop_2_id not in list_all_shops:
        list_all_shops.append(shop_2_id)

print('We have', len(list_all_shops), 'shops in database.')
# Fina a list of suspected shops
array_shop_count = np.zeros(len(list_all_shops))
array_max_rider_count = np.zeros(len(list_all_shops))
print('')
for i in range(0, len(list_shop_1)):
    shop_1_id = list_shop_1[i]
    shop_2_id = list_shop_2[i]
    ind_shop_1 = list_all_shops.index(shop_1_id)
    ind_shop_2 = list_all_shops.index(shop_2_id)
    array_shop_count[ind_shop_1] = array_shop_count[ind_shop_1] + 1
    array_shop_count[ind_shop_2] = array_shop_count[ind_shop_2] + 1
    if list_rider_cnt[i] > array_max_rider_count[ind_shop_1]:
        array_max_rider_count[ind_shop_1] = list_rider_cnt[i]
    if list_rider_cnt[i] > array_max_rider_count[ind_shop_2]:
        array_max_rider_count[ind_shop_2] = list_rider_cnt[i]
# print('Corresponding count: ', array_shop_count)

# Check all suspect shops?
suspect_shop_list = []
for i in range(0, len(list_all_shops)):
    if array_shop_count[i] > 1 and array_max_rider_count[i] > 20:
        # This is a suspect shop
        suspect_shop_list.append(list_all_shops[i])
print('We have',len(suspect_shop_list), 'suspected shop.')

# Suspected shop?
for suspect_shop_id in suspect_shop_list:
    if suspect_shop_id in list_shop_1:
        suspect_lat = list_lat_1[list_shop_1.index(suspect_shop_id)]
        suspect_lon = list_lon_1[list_shop_1.index(suspect_shop_id)]
    else:
        suspect_lat = list_lat_2[list_shop_2.index(suspect_shop_id)]
        suspect_lon = list_lon_2[list_shop_2.index(suspect_shop_id)]

    # Get corresponding shop of a suspected shop
    list_mirror_shops = []
    list_mirror_lat = []
    list_mirror_lon = []
    for i in range(0, len(list_rider_cnt)):
        if list_shop_1[i] == suspect_shop_id:
            # Find corres shops
            list_mirror_shops.append(list_shop_2[i])
            list_mirror_lat.append(list_lat_2[i])
            list_mirror_lon.append(list_lon_2[i])
        if list_shop_2[i] == suspect_shop_id:
            list_mirror_shops.append(list_shop_1[i])
            list_mirror_lat.append(list_lat_1[i])
            list_mirror_lon.append(list_lon_1[i])

    # Plot lat/lon
    gmap = gmplot.GoogleMapPlotter(suspect_lat, suspect_lon, 16)
    # Plot
    gmap.scatter([suspect_lat], [suspect_lon], 'red', marker=True)
    gmap.scatter(list_mirror_lat, list_mirror_lon, 'blue', marker=True)
    file_name = '_'.join(["suspect shop and its mirrors",str(suspect_shop_id),".html"])
    file_path = os.path.join(figure_path, file_name)
    gmap.draw(file_path)
    print("plot done")