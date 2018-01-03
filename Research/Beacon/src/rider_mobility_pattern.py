#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Oct 23 12:00:38 2017

Theme: Rider's mobility pattern

@author: eleme-yi
"""
import gmplot
import datetime
import openpyxl
import numpy as np

rider_id = 10125035
dt = datetime.date(2017,10,22)


rider_trace_file = '_'.join(['data/time_map/rider_trace',str(rider_id)])

wb = openpyxl.load_workbook(rider_trace_file+'.xlsx')
sheet = wb.get_sheet_by_name('Sheet0')

row_count = sheet.max_row
column_count = sheet.max_column

data_num = row_count - 1
lat_all = []
lon_all = []
lat_rst = []
lon_rst = []
lat_cst = []
lon_cst = []


for i in range(0,data_num):    
    shipping_state = sheet.cell(row=i+2,column=4).value
    lat = sheet.cell(row=i+2,column=6).value
    lon = sheet.cell(row=i+2,column=7).value
    lat_all.append(lat)
    lon_all.append(lon)
    if shipping_state == 80:
        lat_rst.append(lat)
        lon_rst.append(lon)
    elif shipping_state == 40:
        lat_cst.append(lat)
        lon_cst.append(lon)
    else:
        raise Exception('Wrong Shipping State')

lat_cen = lat_all[0]
lon_cen = lon_all[0]

sep = '_'
gmap = gmplot.GoogleMapPlotter(lat_cen, lon_cen, 16)
gmap.plot(lat_all, lon_all, 'black', edge_width=10)
gmap.scatter(lat_rst, lon_rst, 'blue', marker=True)
gmap.scatter(lat_cst, lon_cst, 'red', marker=True)
fileName = sep.join(["trace/rider_trace",str(dt),str(rider_id),".html"])
gmap.draw(fileName)