#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jun 12 16:42:32 2017
@author: eleme-yi
Process beacon data in the .xlsx form
"""

import openpyxl
import matplotlib.pyplot as plt
import math
import numpy as np
import json
import gmplot
from collections import Counter
import datetime

from math import radians, cos, sin, asin, sqrt, pow
from scipy.optimize import minimize, rosen, rosen_der

class eval_obj():        
    def __init__(self,date,shop_id,rider_id):
        self.shop_id = shop_id
        self.rider_id = rider_id
        self.date = date

class InputTimeError(Exception):
    def __init__(self,value):
        self.value = value
    def __str__(self):
        return repr(self.value)

# We use this object to map beacon ID to physical location
# We can build a list of beacon_item by reading an excel form
class beacon_item():
    def __init__(self):
        self.uuid = None
        self.major = None
        self.minor = None
        self.latitude = None
        self.longitude = None
        self.floor = None
        self.shop_id = None

class ins_order():
    def __init__(self):
        self.shop_id = None
        self.rider_id = None
        self.date = None
        self.order_id = None
        self.taken_at = None
        self.created_at = None
        self.accept_at = None
        self.arrive_rst_at = None
        self.pickup_at = None
        self.arrive_at = None
        self.taken_arrived_at = None

# Data read from table "dw_tms_interstellar_ins_beacon_data_hour"
class entry():
    def __init__(self):
        self.id = None
        self.uuid = None
        self.user_id = None
        self.rssi = None
        self.major = None
        self.minor = None
        self.date = None
        self.svrHour = None
        self.svrMnut = None
        self.svrScnd = None
        self.crtShop = None
        self.lclHour = None
        self.lclMnut = None
        self.lclScnd = None
        self.latitude = None
        self.longitude = None

# Build a graph of beacons to detect the wrong beacon
class graphBeacon():
    def __init__(self):
        # Note that here we restrict the size of the adj Matrix
        self.size = 100
        self.adjMat = [[0 for i in range(0,self.size)] for i in range(0,self.size)]
        self.minor = [0 for i in range(0,self.size)]
        self.pos = [[0 for i in range(0,2)] for i in range(0,self.size)]
        self.beaconNum = 0
        
    # Find the index of the beacon according to its minor
    def findIndex(self,minor):      
        for i in range(0,len(self.minor)):
            if self.minor[i] == minor:
                return i
        self.minor[self.beaconNum] = minor
        index = self.beaconNum
        self.beaconNum = self.beaconNum + 1
        return index
        
    # Set the position of the beacon, now pos = [latitude, longitude]
    def setPos(self,minor,pos):
        index = self.findIndex(minor)
        self.pos[index] = pos
    
    # Add an edge to the graph
    # An edge between two beacons means the two beacons are detected by the 
    # same device at the same time
    def addEdge(self,minor1,minor2):
        index1 = self.findIndex(minor1)
        index2 = self.findIndex(minor2)
        self.adjMat[index1][index2] = 1
        self.adjMat[index2][index1] = 1
    
    # Note that new beacon node cannot be added once we shrink the graph
    def shrink(self):
        newAdjMat = [[self.adjMat[i][j] for j in range(0,self.beaconNum)] for\
                      i in range(0,self.beaconNum)]
        newPos = [[self.pos[i][j] for j in range(0,2)] for\
                   i in range(0,self.beaconNum)]
        self.adjMat = newAdjMat
        self.pos = newPos
        

# Calculate the time delta in seconds
def time_delta(late_time, erly_time):
    
    return (late_time[0]-erly_time[0])*3600 + (late_time[1]-erly_time[1])*60 \
+late_time[2] - erly_time[2]

def timeAdd(startTime, duration):
    endTime = [0 for i in range(0,3)]
    endTime[0] = startTime[0]
    endTime[1] = startTime[1]
    endTime[2] = startTime[2]
    if startTime[1] + duration < 59:
        endTime[1] = startTime[1] + duration
    else:
        endTime[0] = startTime[0] + 1
        endTime[1] = startTime[1] + duration - 60
    return endTime
        
# Binary signal x is filtered
# Ones lasts less than n times are considered as noise
# Noise are removed
def windowFilter(y,n):
    x = [i for i in y]
    headPos = 0;
    tailPos = 0
    lastNbr = 0;
    for currPos in range(0,len(x)):
        if lastNbr == 0 and x[currPos] == 1:
            headPos = currPos
            if currPos - tailPos < n and headPos <= tailPos:
                for i in range(tailPos,currPos):
                    x[i] = 1
        if lastNbr == 1 and x[currPos] == 0:
            tailPos = currPos
            if currPos - headPos < n and headPos <= tailPos:
                for i in range(headPos,currPos):
                    x[i] = 0
        lastNbr = x[currPos]
    return x

# Read order data from .xlsx file
def readOrderFile(fileName):
    
    wb = openpyxl.load_workbook(fileName+".xlsx")
    sheet = wb.get_sheet_by_name('Sheet0')
    
    row_count = sheet.max_row
    # column_count = sheet.max_column
    
    dataNum = row_count - 1
    
    order_data = []
    
    for i in range(0,dataNum):
        
        order = ins_order()
        
        order.shop_id = sheet.cell(row=i+2,column=2).value
        order.rider_id = sheet.cell(row=i+2,column=19).value
        order.order_id = sheet.cell(row=i+2,column=27).value
                        
        # taken_at time
        dateStr = sheet.cell(row=i+2,column=45).value
        t_a_s = datetime.datetime.strptime(dateStr, "%Y-%m-%d %H:%M:%S.%f")
        order.date = t_a_s.date
        order.taken_at = [t_a_s.hour,t_a_s.minute,t_a_s.second]
        # created_at time
        dateStr = sheet.cell(row=i+2,column=22).value
        t_a_s = datetime.datetime.strptime(dateStr, "%Y-%m-%d %H:%M:%S.%f")
        order.created_at = [t_a_s.hour,t_a_s.minute,t_a_s.second]
        # taken_arrived_at
        dateStr = sheet.cell(row=i+2,column=69).value
        if len(dateStr) == 0:
            print("-- In readOrderFile. Taken_arrived_at missing")
            continue        
        t_a_s = datetime.datetime.strptime(dateStr, "%Y-%m-%d %H:%M:%S.%f")
        order.taken_arrived_at = [t_a_s.hour,t_a_s.minute,t_a_s.second]
        
        order_data.append(order)
                
    return order_data


# Read order data from .xlsx file from Apollo wide detail
def read_talaris_order_file(fileName):
    
    wb = openpyxl.load_workbook(fileName+".xlsx")
    sheet = wb.get_sheet_by_name('Sheet0')
    
    row_count = sheet.max_row
    # column_count = sheet.max_column
    
    dataNum = row_count - 1
    
    order_data = []
    
    for i in range(0,dataNum):
        order_status = sheet.cell(row=i+2,column=10).value
        if order_status != 9:
            continue
        
        order = ins_order()
        
        order.shop_id = sheet.cell(row=i+2,column=39).value
        order.rider_id = sheet.cell(row=i+2,column=46).value
        order.order_id = sheet.cell(row=i+2,column=21).value
        
        # pick_at time
        dateStr = sheet.cell(row=i+2,column=68).value
        t_a_s = datetime.datetime.strptime(dateStr, "%Y-%m-%d %H:%M:%S.%f")
        order.date = t_a_s.date
        order.pickup_at = [t_a_s.hour,t_a_s.minute,t_a_s.second]
        # accept_at time
        dateStr = sheet.cell(row=i+2,column=65).value        
        t_a_s = datetime.datetime.strptime(dateStr, "%Y-%m-%d %H:%M:%S.%f")
        order.accept_at = [t_a_s.hour,t_a_s.minute,t_a_s.second]        
        # accept_rst_at time
        dateStr = sheet.cell(row=i+2,column=67).value
        t_a_s = datetime.datetime.strptime(dateStr, "%Y-%m-%d %H:%M:%S.%f")
        order.arrive_rst_at = [t_a_s.hour,t_a_s.minute,t_a_s.second]        
        # arrive_at time
        dateStr = sheet.cell(row=i+2,column=69).value
        t_a_s = datetime.datetime.strptime(dateStr, "%Y-%m-%d %H:%M:%S.%f")
        order.arrive_at = [t_a_s.hour,t_a_s.minute,t_a_s.second]        
                        
        order_data.append(order)
        
    return order_data

# Read beacon data from talaris file
# Note that the file must be order by time
def beacon2GPS_talaris(fileName):
    
    trace_lat = []
    trace_lon = []
    trace_floor = []
    trace_time = []
    trace_GPS_lat = []
    trace_GPS_lon = []
    
    wb = openpyxl.load_workbook(fileName+".xlsx")
    sheet = wb.get_sheet_by_name('Sheet0') 
    
    row_count = sheet.max_row
    dataNum = row_count - 1
    time_list = []
    obj_list = []        
    
    last_time = None
    num_dup = 0
    for i in range(0,dataNum):
        dateStr = sheet.cell(row=i+2,column=2).value
        if dateStr is None:
            raise Exception("ERROR: NoneType found!!!, i:", i)
            
        t_a_s = datetime.datetime.strptime(dateStr, "%Y-%m-%d %H:%M:%S.%f")
        date = t_a_s.date()
        detail_str = sheet.cell(row=i+2,column=6).value
        if date < datetime.date(2017,9,10):
            obj = json.loads(detail_str)
            detail_obj = obj['location']
            lt = detail_obj['LocationType']
        else:
            detail_obj = json.loads(detail_str)
            lt = detail_obj['location_type']
            
        lclHour = t_a_s.hour
        lclMnut = t_a_s.minute
        lclScnd = t_a_s.second
        time_list.append([lclHour,lclMnut,lclScnd])        
        
        obj_list.append(detail_obj)
        
        hasFound = 0
        #print("lt: ",lt)
        if lt == 10:
            if [lclHour,lclMnut,lclScnd] == last_time:
                num_dup = num_dup + 1
                print(" -- In \"beacon2GPS_talaris\"", num_dup, \
                      " duplicate time data met in talaris file! Ignore!")
                continue
            #print("Found a beacon data.")
            # Append a beacon data
            num_dup = 0
            trace_lat.append(detail_obj['latitude'])
            trace_lon.append(detail_obj['longitude'])
            if date < datetime.date(2017,9,10):
                trace_floor.append(detail_obj['Floor'])
            else:
                trace_floor.append(detail_obj['floor'])
            trace_time.append([lclHour,lclMnut,lclScnd])
            # Find the neareast non-beacon data
            last_time = [lclHour,lclMnut,lclScnd]
            
            j = i
            while j > -1:
                if date < datetime.date(2017,9,10):
                    obj_j_location_type = obj_list[j]['LocationType']
                else:
                    obj_j_location_type = obj_list[j]['location_type']                
                if obj_j_location_type != 10:
                    # A non-beacon type
                    if time_delta(time_list[i],time_list[j]) < 30:
                        # Check if neareast point is less than 30 seconds
                        trace_GPS_lat.append(obj_list[j]['latitude'])
                        trace_GPS_lon.append(obj_list[j]['longitude'])
                        hasFound = 1
                        break
                j = j - 1
            if hasFound == 0:
                # Not found in the excel?
                # raise Exception("Neareast non-beacon data not found!")
                print(" -- In \"beacon2GPS_talaris\" ","At line: ",i+1, \
                      "Corresponding GPS data not found in 30 seconds! Use cached one.")
                if len(trace_GPS_lat) == 0:
                    raise Exception("No cache found!")
                else:
                    trace_GPS_lat.append(trace_GPS_lat[len(trace_GPS_lat)-1])
                    trace_GPS_lon.append(trace_GPS_lon[len(trace_GPS_lon)-1])
        
    return trace_lat, trace_lon, trace_floor, trace_time, trace_GPS_lat, \
trace_GPS_lon
        
# Read beacon data from .xlsx file
def read_raw_talaris_file(fileName,option):
    
    wb = openpyxl.load_workbook(fileName+".xlsx")
    sheet = wb.get_sheet_by_name('Sheet0')
    
    row_count = sheet.max_row
    # column_count = sheet.max_column
    
    
    beacon_data = []
    maxHour = 0
    maxMnut = 0
    maxScnd = 0
    minHour = 23
    minMnut = 59
    minScnd = 59
    
    beacon_at_same_second = []
    num_most_beacon_at_this_second = []
    last_beacon = None
    
    for i in range(0,row_count-3):
        beacon = entry()
        beacon.id = sheet.cell(row=i+2,column=1).value
        beacon.user_id = sheet.cell(row=i+2,column=5).value
        
        beacon_info = sheet.cell(row=i+2,column=6).value
        obj = json.loads(beacon_info)
        beacon.uuid = obj['beaconUuid']
        beacon.rssi = obj['beaconRssi']
        beacon.major = obj['beaconMajor']
        beacon.minor = obj['beaconMinor']
        # Raw data from talaris do not have lat/lon info
        beacon.latitude = None
        beacon.longitude = None

        dateStr = sheet.cell(row=i+2,column=2).value
        if dateStr is None:
            raise Exception("ERROR: NoneType found!!!, i:", i)
            return
        
        # USE STANDARD DATETIME
        lclTime = datetime.datetime.strptime(dateStr, "%Y-%m-%d %H:%M:%S.%f")
        beacon.date = lclTime.date
        beacon.lclHour = lclTime.hour
        beacon.lclMnut = lclTime.minute
        beacon.lclScnd = lclTime.second
        
        if option == "zs":
            if beacon.major == 16160:
                beacon_data.append(beacon)
        elif option == "rs":
            if beacon.major == 16161:
                beacon_data.append(beacon)
        elif option == "all":
            if beacon.uuid == "E2C56DB5-DFFB-48D2-B060-D0F5A71096E0":
                beacon_data.append(beacon)
        else:
            raise Exception("Wrong option!")
        
        
        ## OLD METHODS
        # d1 = dateStr.split(" ")                            
        # beacon_data[i].date = d1[0]
        # time1 = d1[1].split(":")
        # time2 = time1[2].split("+")
        # beacon_data[i].lclHour = int(time1[0])
        # beacon_data[i].lclMnut = int(time1[1])
        # beacon_data[i].lclScnd = int(time2[0])
        
        if i > 0:
            if time_delta([beacon.lclHour, beacon.lclMnut, 
                         beacon.lclScnd], [last_beacon.lclHour, \
                         last_beacon.lclMnut, last_beacon.lclScnd]) == 0:
                # Same second as last entry
                beacon_at_same_second.append(beacon.minor)
            else:
                if len(beacon_at_same_second) > 0:                    
                    num_most_beacon_at_this_second.append(beacon_at_same_second.\
                                                          count(most_common(beacon_at_same_second)))
                beacon_at_same_second = []
                
        
        if time_delta([beacon.lclHour, beacon.lclMnut, 
                     beacon.lclScnd], [maxHour, maxMnut, maxScnd]) > 0:
            maxHour = beacon.lclHour
            maxMnut = beacon.lclMnut
            maxScnd = beacon.lclScnd
        if time_delta([beacon.lclHour, beacon.lclMnut, 
                     beacon.lclScnd], [minHour, minMnut, minScnd]) < 0:
            minHour = beacon.lclHour
            minMnut = beacon.lclMnut
            minScnd = beacon.lclScnd  

        last_beacon = beacon                         
    
    strtTime = [minHour,minMnut,minScnd]
    endTime = [maxHour,maxMnut,maxScnd]
    
    # print("num_most_beacon_at_this_second: ", num_most_beacon_at_this_second)
    if max(num_most_beacon_at_this_second) > 5:
        raise Exception("Duplicate beacon data?! File name: ",fileName)
    
    return beacon_data, strtTime, endTime


# Read beacon data from .xlsx file
def readBeaconFile(fileName):
    
    wb = openpyxl.load_workbook(fileName+".xlsx")
    sheet = wb.get_sheet_by_name('Sheet0')
    
    row_count = sheet.max_row
    # column_count = sheet.max_column
    
    dataNum = row_count - 1
    
    beacon_data = [entry() for i in range(dataNum)]
    maxHour = 0
    maxMnut = 0
    maxScnd = 0
    minHour = 23
    minMnut = 59
    minScnd = 59
    
    
    for i in range(0,dataNum):
        beacon_data[i].id = sheet.cell(row=i+2,column=1).value
        beacon_data[i].user_id = sheet.cell(row=i+2,column=2).value
        beacon_data[i].rssi = sheet.cell(row=i+2,column=9).value
        beacon_data[i].latitude = sheet.cell(row=i+2,column=11).value
        beacon_data[i].longitude = sheet.cell(row=i+2,column=12).value
        beacon_data[i].uuid = sheet.cell(row=i+2,column=13).value
        beacon_data[i].major = int(sheet.cell(row=i+2,column=14).value)
        beacon_data[i].minor = int(sheet.cell(row=i+2,column=15).value)
        dateStr = sheet.cell(row=i+2,column=25).value
        if dateStr is None:
            print("ERROR: NoneType found!!!, i:", i)
            return
        d1 = dateStr.split(" ")                            
        beacon_data[i].date = d1[0]
        time1 = d1[1].split(":")
        time2 = time1[2].split(".")
        beacon_data[i].lclHour = int(time1[0])
        beacon_data[i].lclMnut = int(time1[1])
        beacon_data[i].lclScnd = int(time2[0])
        
        if time_delta([beacon_data[i].lclHour, beacon_data[i].lclMnut, 
                     beacon_data[i].lclScnd], [maxHour, maxMnut, maxScnd]) > 0:
            maxHour = beacon_data[i].lclHour
            maxMnut = beacon_data[i].lclMnut
            maxScnd = beacon_data[i].lclScnd
        if time_delta([beacon_data[i].lclHour, beacon_data[i].lclMnut, 
                     beacon_data[i].lclScnd], [minHour, minMnut, minScnd]) < 0:
            minHour = beacon_data[i].lclHour
            minMnut = beacon_data[i].lclMnut
            minScnd = beacon_data[i].lclScnd                           
        beacon_data[i].crtShop = int(sheet.cell(row=i+2,column=24).value)
    
    strtTime = [minHour,minMnut,minScnd]
    endTime = [maxHour,maxMnut,maxScnd]
    
    return beacon_data, strtTime, endTime
    
# Read the minor-position table and build a list of beacon_item
def ini_beacon_info(fileName):
    
    wb = openpyxl.load_workbook(fileName+".xlsx")
    sheet = wb.get_sheet_by_name('Sheet0')
    
    row_count = sheet.max_row
    # column_count = sheet.max_column
    
    beaconNum = row_count - 1
    
    beacon_info = [beacon_item() for i in range(beaconNum)]
    
    for i in range(0,beaconNum):
        beacon_info[i].shop_id = sheet.cell(row=i+2,column=2).value
        beacon_info[i].uuid = sheet.cell(row=i+2,column=4).value
        beacon_info[i].major = sheet.cell(row=i+2,column=5).value
        beacon_info[i].minor = sheet.cell(row=i+2,column=6).value
        beacon_info[i].floor = sheet.cell(row=i+2,column=9).value
        
        # Original lat/lon
#        beacon_info[i].latitude = sheet.cell(row=i+2,column=7).value
#        beacon_info[i].longitude = sheet.cell(row=i+2,column=8).value
        
        # Revised lat/lon
        beacon_info[i].latitude = sheet.cell(row=i+2,column=12).value
        beacon_info[i].longitude = sheet.cell(row=i+2,column=13).value        
        

    return beacon_info

def talaris_ini_beacon_info(fileName,beacon_info):
    
    wb = openpyxl.load_workbook(fileName+".xlsx")
    sheet = wb.get_sheet_by_name('Sheet0')
    
    row_count = sheet.max_row
    # column_count = sheet.max_column
    
    beaconNum = row_count - 1
    
    # print("beaconNum: ",beaconNum)
    
    for i in range(0,beaconNum):
        major = sheet.cell(row=i+2,column=3).value
        if major == 16160 or major == 16161:
            beacon_info.append(beacon_item())
            index = len(beacon_info)-1
            beacon_info[index].shop_id = sheet.cell(row=i+2,column=7).value
            beacon_info[index].uuid = sheet.cell(row=i+2,column=2).value
            beacon_info[index].major = sheet.cell(row=i+2,column=3).value
            beacon_info[index].minor = sheet.cell(row=i+2,column=4).value
            beacon_info[index].latitude = sheet.cell(row=i+2,column=5).value
            beacon_info[index].longitude = sheet.cell(row=i+2,column=6).value
            beacon_info[index].floor = sheet.cell(row=i+2,column=8).value
    
    return beacon_info
    
# Find the beacon info according to its minor
def find_beacon_by_id(beacon_info,uuid,major,minor):
    for beacon in beacon_info:
        if beacon.uuid == uuid and beacon.major == major \
        and beacon.minor == minor:
            return beacon        
    return -1


# Transfer the rssi value to distance (meters)
# rssi = -65 - 10n*log(d)
# d = 10^((-rssi-65)/10n)
def rssi2distance(rssi): 
    import numpy as np
    # n = 2.8 # Old experiment result
    n = 1.5 # Revised experiment result
    d = np.power(10,(-rssi-65)/(10*n))
    return d

# Read the raw beacon file form the table 
# "dw.dw_tms_interstellar_ins_beacon_data_hour"
# Return the RSSI value regarding to the major, minor and user_id
# The RSSI value is ordered in time sequence with Start time and End time
# The RSSI value is stored per second, if more than one beacon signal is 
# detected, we store the one with highest RSSI level.
def getRssiData(fileName,major,minor,user_id):
    
    beacon_data, strtTime, endTime = readBeaconFile(fileName)
    
    duration = time_delta(endTime,strtTime)+1
    print("duration:",duration)
    
    rssi = [-100 for i in range(0,duration)]
    inShop = [0 for i in range(0,duration)]
    
    # For future use
    #uuid = "8C800121-B6C5-4008-AB08-DB80A138CD6"
    
    shop_id = 1;
    
    for beacon in beacon_data:
        if beacon.major == major and \
            beacon.minor == minor and beacon.user_id == user_id:
                pos = time_delta([beacon.lclHour, beacon.lclMnut, beacon.lclScnd],
                                strtTime)
                rssi[pos] = max(rssi[pos],beacon.rssi)
                if beacon.crtShop == shop_id:
                    inShop[pos] = shop_id

#    plotEnd = len(rssi)
#    plt.clf()    
#    fig = plt.gcf()
#    fig.set_size_inches(10, 8)
#    plt.subplot(2,1,1)
#    plt.plot(rssi[0:plotEnd],'b-')
#    plt.title("RSSI and inShop Plot")
#    plt.subplot(2,1,2)
#    plt.plot(inShop[0:plotEnd],'r-')
#    plt.show()
    
    return rssi, inShop, strtTime, endTime

# Get a specific user's RSSI data from the file (regardless of minors)
def getRssiDataByUser(fileName,user_id,startTime,endTime):
    beacon_data, sTime, eTime = readBeaconFile(fileName)
    
    rider_beacon = []
    #For future use
    uuid = "E2C56DB5-DFFB-48D2-B060-D0F5A71096E0"
    
    for beacon in beacon_data:
        if beacon.uuid == uuid and beacon.user_id == user_id and \
        time_delta([beacon.lclHour ,beacon.lclMnut, beacon.lclScnd],startTime) > 0 and \
                  time_delta([beacon.lclHour ,beacon.lclMnut, beacon.lclScnd],\
                            endTime) < 0:                                           
            rider_beacon.append(beacon)
            
    return rider_beacon

# Get a specific user's RSSI data from the file (regardless of minors)
def getRssiDataByUserByBeacon(fileName,user_id,uuid,major,minor):
    beacon_data, strtTime, endTime = readBeaconFile(fileName)
    
    rider_beacon = []
    #For future use
    uuid = "E2C56DB5-DFFB-48D2-B060-D0F5A71096E0"
    
    for beacon in beacon_data:
        if beacon.uuid == uuid and beacon.user_id == user_id and \
        beacon.major == major and beacon.minor == minor:
            rider_beacon.append(beacon)
            
    return rider_beacon

def get_location_by_shop_id(shop_id):
    
    # print("shop_id need to find: ",shop_id)
    
    beacon_info = ini_beacon_info("restaurant-beacon-mapping")
    beacon_info = talaris_ini_beacon_info("data/others/all_restaurant_beacon_mapping",\
                                          beacon_info)
    if beacon_info is None:
        raise Exception("beacon_info is None?")
    for beacon in beacon_info:
        #print("beacon.shop_id:",beacon.shop_id)
        if beacon.shop_id is None:
            continue
        if beacon.shop_id == shop_id:
            # print("Shop location found!")
            return beacon.latitude, beacon.longitude, beacon.floor
    raise Exception("beacon_info not found!")


def get_beacon_by_major_minor(major,minor,beacon_info):
    for beacon in beacon_info:
        if beacon.minor == minor and beacon.major == major:
            return beacon
    return None
    print("major, minor ", major, minor, "not found! Return None!")
    # raise Exception("beacon not found!")


def get_created_at_from_ins_order_file(order_file, shop_id, rider_id):
    created_at = []
    order_id = []
    order_data = readOrderFile(order_file)
    orderNum = len(order_data)
    # print("orderNum: ",orderNum)
    for i in range(0,orderNum):
        if order_data[i].shop_id == shop_id and order_data[i].rider_id == rider_id:
            created_at.append(order_data[i].created_at)
            order_id.append(order_data[i].order_id)
            
    return created_at, order_id    

def get_timestamp_from_talaris_order_file(order_file, shop_id, rider_id):
    
    accept_at = []
    arrive_rst_at = []
    pickup_at = []
    arrive_at = []
    order_id = []
    
    order_data = read_talaris_order_file(order_file)
    orderNum = len(order_data)
    # print("orderNum: ",orderNum)
    for i in range(0,orderNum):
        if order_data[i].shop_id == shop_id and order_data[i].rider_id == rider_id:
            accept_at.append(order_data[i].accept_at)
            arrive_rst_at.append(order_data[i].arrive_rst_at)
            pickup_at.append(order_data[i].pickup_at)
            arrive_at.append(order_data[i].arrive_at)            
            order_id.append(order_data[i].order_id)
            
    return accept_at, arrive_rst_at, pickup_at, arrive_at, order_id 

def get_timestamp_from_ins_order_file(order_file, shop_id, rider_id):
    
    created_at = []
    taken_at = []
    taken_arrived_at = []
    order_id = []
    
    order_data = readOrderFile(order_file)
    orderNum = len(order_data)
    # print("orderNum: ",orderNum)
    for i in range(0,orderNum):
        if order_data[i].shop_id == shop_id and order_data[i].rider_id == rider_id:
            created_at.append(order_data[i].created_at)
            taken_at.append(order_data[i].taken_at)
            taken_arrived_at.append(order_data[i].taken_arrived_at)
            order_id.append(order_data[i].order_id)
            
    return created_at, taken_at, taken_arrived_at, order_id 


def get_accept_at_from_talaris_order_file(order_file, shop_id, rider_id):
    accept_at = []
    order_id = []
    order_data = read_talaris_order_file(order_file)
    orderNum = len(order_data)
    # print("orderNum: ",orderNum)
    for i in range(0,orderNum):
        if order_data[i].shop_id == shop_id and order_data[i].rider_id == rider_id:
            accept_at.append(order_data[i].accept_at)
            order_id.append(order_data[i].order_id)
            
    return accept_at, order_id 

# Get major and minor from shop id
# Note that major and minor can be a list
def get_major_minor_by_shop_id(shop_id,beacon_info):
    major = []
    minor = []
    for i in range(len(beacon_info)):
        if beacon_info[i].shop_id == shop_id:
            major.append(beacon_info[i].major)
            minor.append(beacon_info[i].minor)
    return major,minor

# This logic is easy, we find the 
def get_arrTime_from_beacon_data(query_time,major,minor,beacon_data):
    # Note that query_time can not be a list here
    if len(query_time) != 3:
        raise Exception("Not single query_time?")
    if len(major) != len(minor):
        raise Exception("major and minor length not equal")
    for i in range(0,len(beacon_data)):
        beacon = beacon_data[i]               
        
        if time_delta(query_time,[beacon.lclHour,beacon.lclMnut,beacon.lclScnd]) > 0:
            # query_time not pass
            continue
        #print("query_time, beacon_time: ",query_time, beacon.lclHour,beacon.lclMnut,beacon.lclScnd)
        for j in range(0,len(major)):
            #print("beacon.major,major,beacon.minor,minor", beacon.major,major,beacon.minor,minor)
            if beacon.major == major[j] and beacon.minor == minor[j]:
                return [beacon.lclHour,beacon.lclMnut,beacon.lclScnd]
        
    #raise Exception("arrTime not found!")
    return -1
    

def get_check_in_location(date,rider_id,shop_id,startTime,endTime,beacon_data,shop_major,shop_minor):
    
    beacon_file = '_'.join(['data/beacon/test_gps',str(date),str(rider_id)])
    beacon_info = ini_beacon_info("restaurant-beacon-mapping")
    rider_beacon = getRssiDataByUser(beacon_file,rider_id,startTime,endTime)
    trace_latitude,trace_longitude,trace_floor,trace_timeStamp,trace_GPS_latitude,\
    trace_GPS_longitude = beacon2GPS(rider_beacon,beacon_info)
    
    order_file = '_'.join(['data/order/ins_order',str(date),str(shop_id)])
    
    # arrTime, order_id = getArrTimeFromInsOrder(order_file, shop_id, rider_id,1)
    # Instead of get take_at time, we should get accept_time
    # For ins order, use created_at time to replace since no accpet_time
    created_at, order_id = get_created_at_from_ins_order_file(order_file, shop_id, rider_id)
    # We use created_at time to find the arrTime
    arrTime = []
    num_check_in_fail = 0
    
    # We also get the riders' reported time for comparison
    rider_arr_time, rider_order_id = getArrTimeFromInsOrder(order_file, shop_id, rider_id,1)
        
    order_num = len(created_at)
    
    for i in range(0,len(created_at)):
        arr_time_entry = get_arrTime_from_beacon_data(created_at[i],\
                                                      shop_major,shop_minor,beacon_data)
        if arr_time_entry == -1:
            print("-- In get_check_in_location. Beacon NOT heard in order shop.")
            print("date, rider_id, shop_id, created_at[i], shop_major, shop_minor",\
                  date, rider_id, shop_id, created_at[i], shop_major, shop_minor)
            num_check_in_fail = num_check_in_fail + 1
            #raise Exception("arr_time not found!")     
            continue

        time_diff = time_delta(rider_arr_time[i],arr_time_entry)
        
        if time_diff > 1800 or time_diff < -1800:
            # Difference is larger than half an hour
            print("-- In talaris_get_check_in_location. Beacon heard but too late or too early.")
            print("date, rider_id, shop_id, created_at[i], shop_major, shop_minor",\
                  date, rider_id, shop_id, created_at[i], shop_major, shop_minor)
            num_check_in_fail = num_check_in_fail + 1
            #raise Exception("arr_time not found!")  
            continue
               
        arrTime.append(arr_time_entry)
    
    punchNum = len(arrTime)
    beacon_lat = []
    beacon_lon = []
    GPS_lat = []
    GPS_lon = []
    arr_order = []
    arr_time = []
    beacon_floor = []
    for i in range(0,punchNum):
        b_lat, b_lon, b_floor, g_lat, g_lon = locationQuery(trace_latitude, \
                 trace_longitude, trace_floor, trace_timeStamp, \
                 trace_GPS_latitude, trace_GPS_longitude, arrTime[i])
        if b_lat == -1:
            continue
        beacon_lat.append(b_lat)
        beacon_lon.append(b_lon)
        beacon_floor.append(b_floor)
        GPS_lat.append(g_lat)
        GPS_lon.append(g_lon)
        arr_order.append(order_id[i])
        arr_time.append(arrTime[i])
    
    if len(beacon_lat) == 0:
        print("No check in found?!")
        #raise Exception("No check in found?!")
        
    
    return beacon_lat,beacon_lon,beacon_floor,GPS_lat,GPS_lon,arr_order,arr_time,order_num,num_check_in_fail


# Get talaris check in using filtering
def ins_get_check_in_location_filter(date,rider_id,shop_id,startTime,\
                                         endTime,beacon_data,shop_major,shop_minor,beacon_info):
    
    trace_lat,trace_lon,trace_floor,trace_time,trace_GPS_lat,trace_GPS_lon = \
    beacon2GPS(beacon_data,beacon_info)
    
    order_file = '_'.join(['data/order/ins_order',str(date),str(shop_id)])
    created_at, taken_at, taken_arrived_at, order_id = \
    get_timestamp_from_ins_order_file(order_file, shop_id, rider_id)    
    
    num_order = len(created_at)
    num_check_in_fail = 0
    
    check_in_dist_beacon = []
    check_in_dist_gps = []
    check_in_time_delta = []
    
    shop_lat, shop_lon, shop_floor = get_location_by_shop_id(shop_id)
    
    for i in range(0,len(created_at)):
        # for each order in the order list
        start_time = created_at[i]
        end_time = taken_arrived_at[i]
        # Find all the rssi values from accepting and delivering
        rssi_list, rssi_time = get_rssi_from_beacon_data(beacon_data,shop_major,\
                                                     shop_minor,start_time,end_time)
        
        if len(rssi_list) == 0:
            print("--In 'ins_get_check_in_location_filter', beacon NOT heard when check in. ")
            print("start_time, end_time, shop_major,shop_minor",\
                  start_time, end_time, shop_major,shop_minor)
            num_check_in_fail = num_check_in_fail + 1
            continue
        
        # Filtering the data with a naive method
        wind_size = 5
        for j in range(wind_size,len(rssi_list)):
            past_rssi = 0
            past_num = 0
            for k in range(0, wind_size):                
                if time_delta(rssi_time[j],rssi_time[j-k]) < wind_size:
                    past_rssi = past_rssi + rssi_list[j-k]
                    past_num = past_num + 1
            rssi_list[j] = past_rssi/past_num                
                            
        # Complement the rssi data and use advanced filtering method
        
        
        # Find max value and calculate distance
        max_rssi = max(rssi_list)
        max_index = rssi_list.index(max_rssi)
        check_in_time = rssi_time[max_index]
        check_in_dist_beacon.append(rssi2distance(max_rssi))
        
        b_lat, b_lon, b_floor, g_lat, g_lon = locationQuery(trace_lat, \
                 trace_lon, trace_floor, trace_time, \
                 trace_GPS_lat, trace_GPS_lon, rssi_time[max_index])
        if b_lat == -1:
            raise Exception("Beacon heard but no GPS data?!")
        gps_lat_delta_sqr = (shop_lat - g_lat)*(shop_lat - g_lat)
        gps_lon_delta_sqr = (shop_lon - g_lon)*(shop_lon - g_lon)
        gps_dist = (math.sqrt(gps_lat_delta_sqr+gps_lon_delta_sqr))*100000
        check_in_dist_gps.append(gps_dist)
        check_in_time_delta.append(abs(time_delta(check_in_time,taken_at[i])))
    
    return check_in_dist_beacon,check_in_dist_gps,check_in_time_delta,num_order,num_check_in_fail


# Get rssi list from beacon_data by major, minor, and start_time, end_time
# Note that major and minor can be list
def get_rssi_from_beacon_data(beacon_data,major,minor,start_time,end_time):
    rssi_list = []
    rssi_time = []
    for beacon in beacon_data:
        beacon_time = [beacon.lclHour,beacon.lclMnut,beacon.lclScnd]
        if time_delta(start_time,beacon_time) > 0 or time_delta(beacon_time,end_time) > 0:
            # if not in the interval
            continue
        for i in range(0,len(major)):
            #print("beacon_time,beacon.major,beacon.minor",beacon_time,beacon.major,beacon.minor)
            if beacon.major == major[i] and beacon.minor == minor[i]:
                rssi_list.append(beacon.rssi)
                rssi_time.append(beacon_time)
                break            
            
    if len(rssi_list) == 0:
        print("rssi not found in the time interval")
    
    return rssi_list, rssi_time


# Get talaris check in using filtering
def talaris_get_check_in_location_filter(date,rider_id,shop_id,startTime,\
                                         endTime,beacon_data,shop_major,shop_minor):
    
    talaris_file = '_'.join(['data/talaris/all_talaris',str(date),str(rider_id)])
    trace_lat,trace_lon,trace_floor,trace_time,trace_GPS_lat,trace_GPS_lon = \
    beacon2GPS_talaris(talaris_file)
    
    order_file = '_'.join(['data/order/talaris_order',str(date),str(shop_id)])    
    accept_at, arrive_rst_at, pickup_at, arrive_at, order_id = \
    get_timestamp_from_talaris_order_file(order_file,shop_id,rider_id)
    
    num_order = len(accept_at)
    num_check_in_fail = 0
    
    check_in_dist_beacon = []
    check_in_dist_gps = []
    check_in_time_delta = []
    GPS_lat = []
    GPS_lon = []
    
    shop_lat, shop_lon, shop_floor = get_location_by_shop_id(shop_id)
    
    for i in range(0,len(accept_at)):
        # for each order in the order list
        start_time = accept_at[i]
        end_time = arrive_at[i]
        # Find all the rssi values from accepting and delivering
        rssi_list, rssi_time = get_rssi_from_beacon_data(beacon_data,shop_major,\
                                                     shop_minor,start_time,end_time)
        
        if len(rssi_list) == 0:
            print("--In 'talaris_get_check_in_location_filter', beacon NOT heard when check in. ")
            print("start_time, end_time, shop_major,shop_minor",\
                  start_time, end_time, shop_major,shop_minor)
            num_check_in_fail = num_check_in_fail + 1
            continue
        
        # Filtering the data with a naive method
        wind_size = 5
        for j in range(wind_size,len(rssi_list)):
            past_rssi = 0
            past_num = 0
            for k in range(0, wind_size):                
                if time_delta(rssi_time[j],rssi_time[j-k]) < wind_size:
                    past_rssi = past_rssi + rssi_list[j-k]
                    past_num = past_num + 1
            rssi_list[j] = past_rssi/past_num                
                            
        # Complement the rssi data and use advanced filtering method
        
        
        # Find max value and calculate distance
        max_rssi = max(rssi_list)
        max_index = rssi_list.index(max_rssi)
        check_in_time = rssi_time[max_index]
        check_in_dist_beacon.append(rssi2distance(max_rssi))
        
        b_lat, b_lon, b_floor, g_lat, g_lon = locationQuery(trace_lat, \
                 trace_lon, trace_floor, trace_time, \
                 trace_GPS_lat, trace_GPS_lon, rssi_time[max_index])
        if b_lat == -1:
            raise Exception("Beacon heard but no GPS data?!")
            
        GPS_lat.append(g_lat)
        GPS_lon.append(g_lon)
        gps_lat_delta_sqr = (shop_lat - g_lat)*(shop_lat - g_lat)
        gps_lon_delta_sqr = (shop_lon - g_lon)*(shop_lon - g_lon)
        gps_dist = (math.sqrt(gps_lat_delta_sqr+gps_lon_delta_sqr))*100000
        check_in_dist_gps.append(gps_dist)
        check_in_time_delta.append(abs(time_delta(check_in_time,arrive_rst_at[i])))
    
    
    # Draw check in place
    center_lat = shop_lat
    center_lon = shop_lon
    
    gmap = gmplot.GoogleMapPlotter(center_lat, center_lon, 16)
    gmap.scatter(GPS_lat, GPS_lon, 'red', size=20, marker=True)
    gmap.scatter([shop_lat], [shop_lon],'black', size=60, marker=True)
    figureName = '_'.join(['trace/arrival/arrival_comp',str(date),str(shop_id),str(rider_id)])
    figureFile = '.'.join([figureName,'html'])
    gmap.draw(figureFile)
    print("plot done")
    
    
    
    return check_in_dist_beacon,check_in_dist_gps,check_in_time_delta,num_order,num_check_in_fail

# Same function as "get_check_in_location"
# Rewrite for reading talaris related files
def talaris_get_check_in_location(date,rider_id,shop_id,startTime,endTime,beacon_data,shop_major,shop_minor):
    talaris_file = '_'.join(['data/talaris/all_talaris',str(date),str(rider_id)])
    trace_lat,trace_lon,trace_floor,trace_time,trace_GPS_lat,trace_GPS_lon = \
    beacon2GPS_talaris(talaris_file)
    
    order_file = '_'.join(['data/order/talaris_order',str(date),str(shop_id)])
    #arrTime, order_id = getArrTimeFromInsOrder(order_file, shop_id, rider_id,2)
    # Instead of get take_at time, we should get accept_time
    # For ins order, use created_at time to replace since no accpet_time
    accept_at, order_id = get_accept_at_from_talaris_order_file(order_file, shop_id, rider_id)
    # We use created_at time to find the arrTime    
    arrTime = []
    
    # We also get the riders' reported time for comparison
    rider_arr_time, rider_order_id = getArrTimeFromInsOrder(order_file, shop_id, rider_id,2)
    
    order_num = len(accept_at)
    
    num_check_in_fail = 0
    for i in range(0,len(accept_at)):
        arr_time_entry = get_arrTime_from_beacon_data(accept_at[i],\
                                               shop_major,shop_minor,beacon_data)                        
        if arr_time_entry == -1:
            print("-- In talaris_get_check_in_location. Beacon NOT heard in order shop.")
            print("date, rider_id, shop_id, accept_at[i], shop_major, shop_minor",\
                  date, rider_id, shop_id, accept_at[i], shop_major, shop_minor)
            num_check_in_fail = num_check_in_fail + 1
            #raise Exception("arr_time not found!")  
            continue

        time_diff = time_delta(rider_arr_time[i],arr_time_entry)
        
        if time_diff > 1800 or time_diff < -1800:
            # Difference is larger than half an hour
            print("-- In talaris_get_check_in_location. Beacon heard but too late or too early.")
            print("date, rider_id, shop_id, accept_at[i], shop_major, shop_minor",\
                  date, rider_id, shop_id, accept_at[i], shop_major, shop_minor)
            num_check_in_fail = num_check_in_fail + 1
            #raise Exception("arr_time not found!")  
            continue
        
        print("Successful beacon info found! shop_major, shop_minor: ",\
              shop_major, shop_minor)
        print("order_id[i], accept_at[i], rider_arr_time[i], arr_time_entry: "\
              ,order_id[i], accept_at[i], rider_arr_time[i], arr_time_entry)
        
        arrTime.append(arr_time_entry)

    punchNum = len(arrTime)
    beacon_lat = []
    beacon_lon = []
    GPS_lat = []
    GPS_lon = []
    arr_order = []
    arr_time = []
    beacon_floor = []
    for i in range(0,punchNum):
        b_lat, b_lon, b_floor, g_lat, g_lon = locationQuery(trace_lat, \
                 trace_lon, trace_floor, trace_time, \
                 trace_GPS_lat, trace_GPS_lon, arrTime[i])
        if b_lat == -1:
            continue
        beacon_lat.append(b_lat)
        beacon_lon.append(b_lon)
        beacon_floor.append(b_floor)
        GPS_lat.append(g_lat)
        GPS_lon.append(g_lon)
        arr_order.append(order_id[i])
        arr_time.append(arrTime[i])
    
    return beacon_lat,beacon_lon,beacon_floor,GPS_lat,GPS_lon,arr_order,arr_time,order_num,num_check_in_fail

def arrivalDetect(raw_rider_beacon):
    
    # Here we first pre-process the data
    rawBeaconNum = len(raw_rider_beacon)
    rider_beacon = []
    for i in range(1,rawBeaconNum):
        if raw_rider_beacon[i].lclScnd != raw_rider_beacon[i-1].lclScnd:
            rider_beacon.append(raw_rider_beacon[i])
        elif raw_rider_beacon[i].rssi > raw_rider_beacon[i-1].rssi:
            rider_beacon[len(rider_beacon)-1].rssi = raw_rider_beacon[i].rssi
    
    windSize = 5
    timeThresh = 10
    beaconNum = len(rider_beacon)
    isInShop = []
    timeStamp = []
    for i in range(windSize,beaconNum):
#        startHour = rider_beacon[i-windSize].lclHour
#        startMnut = rider_beacon[i-windSize].lclMnut
#        startScnd = rider_beacon[i-windSize].lclScnd
#        endHour = rider_beacon[i].lclHour
#        endMnut = rider_beacon[i].lclMnut
#        endScnd = rider_beacon[i].lclScnd
#        lastHour = rider_beacon[i-1].lclHour
#        lastMnut = rider_beacon[i-1].lclMnut
#        lastScnd = rider_beacon[i-1].lclScnd
#        if time_delta([endHour,endMnut,endScnd],[lastHour,lastMnut,lastScnd]) == 0:
#            continue
#        endTime = [endHour,endMnut,endScnd]
#        timeStamp.append(endTime)
#        if time_delta([endHour,endMnut,endScnd],[startHour,startMnut,startScnd]) < \
#        timeThresh:
#            isInShop.append(1)
#        else:
#            isInShop.append(0)


        # A very simple implementation
        timeStamp.append([rider_beacon[i].lclHour,rider_beacon[i].lclMnut,\
                          rider_beacon[i].lclScnd])
        if rider_beacon[i].rssi > -90:
            isInShop.append(1)
        else:
            isInShop.append(0)   
            
    isInShop = windowFilter(isInShop,20)
    
    return isInShop, timeStamp
        

# Get simultaneous beacon data
def get_simul_beacon_data(rider_beacon,index):
    if index < len(rider_beacon):
        simul_beacon_data = [rider_beacon[index]]
        timeStamp = [rider_beacon[index].lclHour,rider_beacon[index].lclMnut,\
                     rider_beacon[index].lclScnd]
        index = index + 1        
        while index < len(rider_beacon) and \
            rider_beacon[index].lclHour == rider_beacon[index-1].lclHour and \
            rider_beacon[index].lclMnut == rider_beacon[index-1].lclMnut and \
            rider_beacon[index].lclScnd == rider_beacon[index-1].lclScnd:
                simul_beacon_data.append(rider_beacon[index])
                index = index + 1        
    else:
        timeStamp = []
        simul_beacon_data = []
        index = index + 1
    
    return simul_beacon_data, index, timeStamp


# Get subsequent beacon data
def getSubBeaconData(rider_beacon,index):
    ini_index = index
    ely_index = index
    if index < len(rider_beacon):
        sub_beacon_data = [rider_beacon[index]]
        timeStamp = [rider_beacon[index].lclHour,rider_beacon[index].lclMnut,\
                     rider_beacon[index].lclScnd]
        index = index + 1        
        while index < len(rider_beacon) and \
            rider_beacon[index].lclHour == rider_beacon[index-1].lclHour and \
            rider_beacon[index].lclMnut == rider_beacon[index-1].lclMnut and \
            rider_beacon[index].lclScnd == rider_beacon[index-1].lclScnd:
                sub_beacon_data.append(rider_beacon[index])
                index = index + 1        
    else:
        timeStamp = []
        sub_beacon_data = []
        index = index + 1
        
        
    # Also attach some early time data for smoothing
    windSize = 10    # We get beacon data in the past and future 3 second
    if ini_index < len(rider_beacon):     
        while ely_index > 0 and \
        time_delta([rider_beacon[ini_index].lclHour,
                  rider_beacon[ini_index].lclMnut, \
                  rider_beacon[ini_index].lclScnd], \
                  [rider_beacon[ely_index].lclHour, \
                  rider_beacon[ely_index].lclMnut, \
                  rider_beacon[ely_index].lclScnd]) < windSize:
            sub_beacon_data.append(rider_beacon[ely_index])
            ely_index = ely_index - 1
                
    return sub_beacon_data, index, timeStamp


#def great_circle_distance(lat1, lon1, lat2, lon2):
#
#    """
#    Calculate the great circle distance between two points 
#    on the earth (specified in decimal degrees)
#    """
#    # convert decimal degrees to radians 
#    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
#
#    # haversine formula 
#    dlon = lon2 - lon1 
#    dlat = lat2 - lat1 
#    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
#    c = 2 * asin(sqrt(a)) 
#    r = 6371 # Radius of earth in kilometers. Use 3956 for miles
#    return c * r

def great_circle_distance(lat1, lon1, lat2, lon2):
    # Note that this function only works in Shanghai
    dist = calDist([lat1,lon1],[lat2,lon2])
    dist = dist * 100000
    # if dist > 100:
    #     print('lat1, lon1, lat2, lon2: ',lat1, lon1, lat2, lon2)
    return dist

# Mean Square Error
# locations: [ (lat1, long1), ... ]
# distances: [ distance1, ... ]
def mse(x, locations, distances):
    mse = 0.0
    for location, distance in zip(locations, distances):
        distance_calculated = great_circle_distance(x[0], x[1], location[0], location[1])
        mse += pow(distance_calculated - distance, 2.0)
        #print("distance_calculated in mse:",distance_calculated)
    #print("mse / len(distances): ",mse / len(distances))
    return mse / len(distances)

# Calculate the device position based on the some beacon RSSI data
# beaconNum is the number of beacons detected
# pos can be vector a GPS or x-y coordinates
def beaconCalPos(raw_rssi, raw_latitude, raw_longitude, raw_floor, pastTime, \
                 method):
    
    if method != 2 and method != 3:
        raise Exception('Wrong method option')
    timeThresh = 5
    pastThresh = -5
    rawNum = len(raw_rssi)
    rssi = []
    latitude = []
    longitude = []
    floor = []
    if pastTime[0] < pastThresh:
        raise Exception('The first time in pastTime is too early.')
    for i in range(0, rawNum):
        if pastTime[i] > -timeThresh:
            rssi.append(raw_rssi[i])
            latitude.append(raw_latitude[i])
            longitude.append(raw_longitude[i])
            floor.append(raw_floor[i])
    if len(rssi) == 0:
        print('rawNum: ',rawNum)
        print('raw_rssi: ',raw_rssi)
        print('pastTime: ',pastTime)
        #raise Exception('Input pastTime is error')
        raise InputTimeError('Input pastTime is error')
    
    if None in rssi:
        raise Exception('None in rssi')
    if None in latitude or None in longitude:
        raise Exception('None in latitude/longitude')
    beaconNum = len(rssi)
    if beaconNum != len(latitude) or beaconNum != len(longitude):
        raise Exception('rssi and latitude not same size')
    w = [0.01 for i in range(0,beaconNum)]
    modRssi = [i+100 for i in rssi]
    
    # A very first version, we use RSSI as weight
    # w = modRssi
    
    # Second version, distance is used as weight
    if method == 2:
        alpha = 1
        w = [math.exp(alpha*i) for i in modRssi]    
        if any(n < 0 for n in w):
            raise Exception('Negative values in weight')    
        totW = np.sum(w)
        lat_sum = 0
        lon_sum = 0
        for i in range(0, beaconNum):
            lat_sum = lat_sum + latitude[i]*w[i]
            lon_sum = lon_sum + longitude[i]*w[i]
        dev_lat = lat_sum/totW
        dev_lon = lon_sum/totW

    # Third version, trilateration
    # Initial point: the point with the closest distance
    if method == 3:
        distances = []
        locations = []
        for i in range(0,len(rssi)):
            distances.append(rssi2distance(rssi[i]))
            locations.append((latitude[i],longitude[i]))

        print("distances:",distances)
        
        min_distance     = float('inf')
        closest_location = None
        for i in range(0,len(rssi)):
            # A new closest point!
            if distances[i] < min_distance:
                min_distance = distances[i]
                closest_location = locations[i]
        initial_location = closest_location
        print("initial_location: ",initial_location)
        
        initial_mse = mse(initial_location, locations, distances)
        print("initial_mse: ",initial_mse)
        
        # initial_location: (lat, long)
        # locations: [ (lat1, long1), ... ]
        # distances: [ distance1,     ... ] 
        
        result = minimize(
            mse,                         # The error function
            initial_location,            # The initial guess
            args=(locations, distances), # Additional parameters for mse
            method='Nelder-Mead',           # The optimisation algorithm
            options={
                # 'ftol':1e-3,         # Tolerance
                'disp': True,
                # 'gtol': 1e-4,
                # 'eps': 1e-4,
                # 'maxiter': 1e+7      # Maximum iterations
            })
        location = result.x
        dev_lat = location[0]
        dev_lon = location[1]
        final_mse = mse(location, locations, distances)
        print("final_mse: ",final_mse)
        print("result.nit: ",result.nit)
        print("result.status",result.status)
        print("result.message",result.message)
        print("result.success: ",result.success)
        
    
    # We use the most common floor as the device's floor
    dev_floor = most_common(floor)
    
    if math.isnan(dev_lat) or math.isnan(dev_lon):
        # print('NaN error:')
        # print('w: ',w, 'totW: ',totW, 'latitude: ',latitude, 'longitude: ', \
        #     longitude)
        raise Exception('NaN in calculated GPS')
    
    return dev_lat, dev_lon, dev_floor

    
# Get the position of beacon by reverse-localization
def getBeaconPos(fileName,rider_id,uuid,major,minor,startTime,endTime,method):
        
    # Initial beacon info and remove if the beacon is known
    foundBeacon = 0
    beacon_info = ini_beacon_info("restaurant-beacon-mapping")
    for i in range(0,len(beacon_info)):
        if beacon_info[i].uuid == uuid and beacon_info[i].major == major and \
        beacon_info[i].minor == minor:
            real_lat = beacon_info[i].latitude
            real_lon = beacon_info[i].longitude
            foundBeacon = 1
            del beacon_info[i]
            break
            
    if foundBeacon == 0:
        raise Exception("Beacon not found in the map?")
            
    # Get the position information of the user
    rider_beacon = getRssiDataByUser(fileName,rider_id,startTime,endTime)
    trace_latitude_2,trace_longitude_2,trace_latitude_3,trace_longitude_3,\
    trace_floor,trace_timeStamp,trace_GPS_latitude,\
    trace_GPS_longitude = beacon2GPS(rider_beacon,beacon_info)
    
    # Get the target beacon info
    beacon_unknown = getRssiDataByUserByBeacon(fileName,rider_id,uuid,major,minor)

    # Prepare variables
    rssi = []
    latitude = []
    longitude = []
    floor = []
    pastTime = []

    # Match the unknown beacon info and the rider info
    index_rider = 0
    index_beacon = 0    
    while index_rider < len(trace_timeStamp) and index_beacon < len(beacon_unknown):
        time_rider = trace_timeStamp[index_rider]
        time_beacon = [beacon_unknown[index_beacon].lclHour,\
                       beacon_unknown[index_beacon].lclMnut,\
                       beacon_unknown[index_beacon].lclScnd]
        if time_rider[0] == time_beacon[0] and time_rider[1] == time_beacon[1] \
        and time_rider[2] == time_beacon[2]:
            # Time match! Attach the data
            rssi.append(beacon_unknown[index_beacon].rssi)
            if method == 2:
                latitude.append(trace_latitude_2[index_rider])
                longitude.append(trace_longitude_2[index_rider])
            elif method == 3:
                latitude.append(trace_latitude_3[index_rider])
                longitude.append(trace_longitude_3[index_rider])
            else:
                raise Exception("Not valid method choice")
            floor.append(trace_floor[index_rider])
            pastTime.append(0)
        if time_delta(time_rider,time_beacon) > 0:
            index_beacon = index_beacon + 1
        else:
            index_rider = index_rider + 1
            
    print("beacon position calculated based on ",len(rssi), "records")
    try:
        beacon_lat, beacon_lon, beacon_floor = beaconCalPos\
        (rssi, latitude, longitude, floor, pastTime, method)
    except InputTimeError:
        print("Oops! InputTimeError! Input time of the beacon is too early?")
    
    return beacon_lat, beacon_lon, beacon_floor, real_lat, real_lon
    
    
# Build a beacon-relation-graph
def buildGraph(rider_beacon,beacon_info):
    beacon_graph = graphBeacon()
    index = 0
    while index < len(rider_beacon):
        sub_beacon_data, index, timeStamp = getSubBeaconData(rider_beacon,index)
        beaconNum = len(sub_beacon_data)
        minor_list = []
        for i in range(0,beaconNum):
            uuid = sub_beacon_data[i].uuid
            major = sub_beacon_data[i].major
            minor = sub_beacon_data[i].minor
            beacon = find_beacon_by_id(beacon_info,uuid,major,minor)
            if beacon == -1:
                continue
            # Ignore the beacon data if we don't know its location
            if beacon.latitude is None or beacon.longitude is None:
                continue
            minor_list.append(minor)
            beacon_graph.setPos(minor,[beacon.latitude,beacon.longitude])
        for minor1 in minor_list:
            for minor2 in minor_list:
                if minor1 != minor2:
                    beacon_graph.addEdge(minor1,minor2)
    return beacon_graph
            

# Calculate the Euclidean distance
def calDist(pos1,pos2):
    import math
    lat_delta = pos1[0] - pos2[0]
    lon_delta = pos1[1] - pos2[1]
    dist = math.sqrt(lat_delta*lat_delta + lon_delta*lon_delta)
    return dist

# Find the wrong placed beacon according to the beacon graph
def findWrongBeacon(beacon_graph):
    beacon_graph.shrink()
    oldGraph = [[0 for i in range(0,beacon_graph.beaconNum)] \
                 for i in range(0,beacon_graph.beaconNum)]
    for i in range(0,beacon_graph.beaconNum):
        for j in range(0,beacon_graph.beaconNum):
            oldGraph[i][j] = beacon_graph.adjMat[i][j]
    for i in range(0,beacon_graph.beaconNum):
        for j in range(0,beacon_graph.beaconNum):
            if i == j:
                continue
            if beacon_graph.adjMat[i][j] == 1:
                dist = calDist(beacon_graph.pos[i],beacon_graph.pos[j])
                # print("pos1: ",beacon_graph.pos[i], \
                #      "pos2: ",beacon_graph.pos[j], "dist: ", dist)
                if dist > 0.001:     # pos 0.001 equals 100 meters here
                    beacon_graph.adjMat[i][j] = 0
                    beacon_graph.adjMat[j][i] = 0
    newGraph = beacon_graph.adjMat
    delta_Graph = [[oldGraph[i][j]-newGraph[i][j] for j in \
                    range(0,beacon_graph.beaconNum)] for i in \
            range(0,beacon_graph.beaconNum)]
    print("oldGraph: ", oldGraph)
    print("newGraph: ", newGraph)
    print("delta_Graph: ", delta_Graph)
                

# Output the trace of the rider according to the rider's beacon record
def beacon2GPS(rider_beacon,beacon_info):
    index = 0
    trace_latitude_2 = []
    trace_longitude_2 = []
    trace_latitude_3 = []
    trace_longitude_3 = []
    trace_timeStamp = []
    trace_floor = []
    trace_GPS_latitude = []
    trace_GPS_longitude = []
    last_lat = None
    last_lon = None
    last_timeStamp = None
    
    while index < len(rider_beacon):
        # print("index in beacon2GPS: ", index)
        sub_beacon_data, index, timeStamp = getSubBeaconData(rider_beacon,index)
        beaconNum = len(sub_beacon_data)
        rssi = []
        latitude = []
        longitude = []
        floor = []
        pastTime = []
        for i in range(0,beaconNum):
            uuid = sub_beacon_data[i].uuid
            major = sub_beacon_data[i].major
            minor = sub_beacon_data[i].minor
            beacon = find_beacon_by_id(beacon_info,uuid,major,minor)
            if beacon == -1:
                continue
            # Ignore the beacon data if we don't know its location
            if beacon.latitude is None or beacon.longitude is None:
                continue
            rssi.append(sub_beacon_data[i].rssi)
            latitude.append(beacon.latitude)
            longitude.append(beacon.longitude)
            floor.append(beacon.floor)
            delta_time = time_delta([sub_beacon_data[i].lclHour,\
                                   sub_beacon_data[i].lclMnut,\
                                   sub_beacon_data[i].lclScnd],\
                                   timeStamp)
            pastTime.append(delta_time)
        if len(rssi) == 0:
            #print("We jump in beacon2GPS! because len(rssi) == 0")
            continue
        if pastTime[0] <= -5:
            #print(pastTime)
            #print("We jump in beacon2GPS! because the first timeStamp in pastTime is too early.")
            continue
        try:
            dev_lat_2, dev_lon_2, dev_floor = beaconCalPos(rssi,latitude,\
                                                           longitude,floor,\
                                                           pastTime,2)
        except InputTimeError:
            print("InputTimeError")
            print(pastTime)
            continue
#        dev_lat_3, dev_lon_3, dev_floor = beaconCalPos(rssi,latitude,longitude,\
#                                                   floor,pastTime,3)
        
        # Here we check and correct the feasibility of current position by 
        # comparing with the last location
#        vThresh = 8     # Threshold is set as 8 m/s        
#        if last_lat != None:
#            delta_time = time_delta(timeStamp,last_timeStamp)
#            # Distance is transferred to meters 
#            delta_dist = calDist([dev_lat,dev_lon],[last_lat,last_lon])*100000
#            # If too quick
#            if (delta_dist/delta_time) > vThresh:
#                ratio = (delta_dist/delta_time)/vThresh
#                dev_lat = last_lat + (dev_lat-last_lat)/ratio
#                dev_lon = last_lon + (dev_lon-last_lon)/ratio 
#
#        last_lat = dev_lat
#        last_lon = dev_lon
#        last_timeStamp = timeStamp   
#        # Append the new location
#        # print("Append new location in beacon2GPS!")
        
        trace_latitude_2.append(dev_lat_2)
        trace_longitude_2.append(dev_lon_2)
        # trace_latitude_3.append(dev_lat_3)
        # trace_longitude_3.append(dev_lon_3)
        
        trace_timeStamp.append(timeStamp)
        trace_floor.append(dev_floor)
        trace_GPS_latitude.append(sub_beacon_data[0].latitude)
        trace_GPS_longitude.append(sub_beacon_data[0].longitude)
    
#    return trace_latitude_2, trace_longitude_2, trace_latitude_3, \
#trace_longitude_3, trace_floor, trace_timeStamp, \
#trace_GPS_latitude, trace_GPS_longitude

    return trace_latitude_2, trace_longitude_2, trace_floor, trace_timeStamp, \
trace_GPS_latitude, trace_GPS_longitude 

    
# Query the beacon-based and GPS-based location    
def locationQuery(trace_latitude, trace_longitude, trace_floor, trace_timeStamp, \
trace_GPS_latitude, trace_GPS_longitude, queryTime):
    
    hasFound = 0
    for i in range(0,len(trace_latitude)-1):
        
        if  type(queryTime) == int or type(trace_timeStamp[i]) == int:
            print("queryTime: ",queryTime)
            print("trace_timeStamp[i]: ",trace_timeStamp[i])
            raise Exception("Invalid time format.")
        
        frstDelta = time_delta(trace_timeStamp[i], queryTime)
        scndDelta = time_delta(trace_timeStamp[i+1], queryTime) 
        #print("queryTime: ",queryTime)
        #print("trace_timeStamp[i]: ",trace_timeStamp[i])
        #print("trace_timeStamp[i+1]: ",trace_timeStamp[i+1])
        #print("frstDelta: ", frstDelta)
        #print("scndDelta: ", scndDelta)
        if frstDelta <= 0 and scndDelta >= 0:
            if frstDelta + scndDelta <= 0:
                index = i+1
            else:
                index = i
            currLat = trace_latitude[index]
            currLon = trace_longitude[index]
            GPSlat = trace_GPS_latitude[index]
            GPSlon = trace_GPS_longitude[index]
            currFloor = trace_floor[index]
            hasFound = 1
            if GPSlat == 0:
                #raise Exception("Invalid GPS data.")
                print("--GPS lat = 0 found in locationQuery")
                hasFound = 0
            break
    
    if hasFound == 0:
        # raise NameError('No match timeStamp')
        # print('No match timeStamp. Due to not enough beacon data')
        return -1, -1, -1, -1, -1
        
                                                
    return currLat, currLon, currFloor, GPSlat, GPSlon
    

def getArrTimeFromInsOrder(fileName, shop_id, rider_id, file_type):
    if file_type == 1:
        order_data = readOrderFile(fileName)
    elif file_type == 2:
        order_data = read_talaris_order_file(fileName)
    else:
        raise Exception("Wrong file_type!")
    arrTime = []
    order_id = []
    orderNum = len(order_data)
    # print("orderNum: ",orderNum)
    for i in range(0,orderNum):
        if order_data[i].shop_id == shop_id and order_data[i].rider_id == rider_id:
            arrTime.append(order_data[i].taken_at)
            order_id.append(order_data[i].order_id)
            
    return arrTime, order_id
    

# A naive method for industry application
# For a rssi sequence, the corresponding rider is considered to enter the shop
# area if rssi signal is detected for more than 10 seconds
def detArrDptVer1(rssi, inShop, strtTime, endTime):
    
    # Step 1: Filter the RSSI signal with a low pass filter
    from scipy import signal
    fs = 1
    nyq = 0.5 * fs
    wn = 0.05 / nyq
    order = 6
    b, a = signal.butter(order, wn, 'low',analog = False)
    filtRssi = signal.filtfilt(b, a, rssi)
    
    # Step 2: Compute inRegion index
    inRegion = [0 for i in range(0,len(filtRssi))]
    for i in range(0,len(filtRssi)):
        if filtRssi[i] > -90:
            inRegion[i] = 1
    print("In Region before window: ",inRegion)
    
    # Step 3: Remove the peaks
    windSize = 20
    filtInRegion = windowFilter(inRegion,windSize)
    print("In Region after window: ",inRegion)
    
    # Plot and compare
    # plotEnd = len(rssi)   # If we want to plot the full data
    plotEnd = 400           # If we only want to plot partial data
    plt.clf()    
    fig = plt.gcf()
    fig.set_size_inches(10, 8)
    plt.subplot(5,1,1)
    plt.plot(rssi[0:plotEnd],'b-',label = "Raw RSSI")
    plt.legend()
    plt.ylabel("dB")
    plt.title("From RSSI to In-Region determination")
    plt.subplot(5,1,2)
    plt.plot(filtRssi[0:plotEnd],'r-',label = "Filtered RSSI")
    plt.legend()
    plt.ylabel("dB")
    plt.subplot(5,1,3)
    plt.plot(inRegion[0:plotEnd],'g-',label = "Initial In-Region")
    plt.legend()
    plt.subplot(5,1,4)
    plt.plot(filtInRegion[0:plotEnd],'c-',label = "Shaved In-Region")
    plt.legend()
    plt.subplot(5,1,5)
    plt.plot(inShop[0:plotEnd],'k-',label = "Real In-Region")
    plt.legend()
    plt.xlabel("Seconds")
    plt.show()
    
def most_common(lst):
    return max(set(lst), key=lst.count)

def two_permu(lst):
    permu = []
    lst_len = len(lst)
    for i in range(0,lst_len):
        for j in range(i+1,lst_len):
            if lst[i] == lst[j]:
                raise Exception("Two number in the list should not be the same")
            if lst[i] < lst[j]:
                permu.append([lst[i],lst[j]])
            else:
                permu.append([lst[j],lst[i]])
    return permu

# calculate the similarity index of two beacons according to the number of 
# times that they are detected at the same time
def cal_sim_index(fileName,test_beacons):
    beacon_data, strtTime, endTime = readBeaconFile(fileName)
    beaconNum = len(test_beacons)
    sim_mat = [[0 for i in range(beaconNum)] for j in range(beaconNum)]
    sco_mat = [[0 for i in range(beaconNum)] for j in range(beaconNum)]
    
    index = 0
    while index < len(beacon_data):
        simul_beacon_data, index, timeStamp = \
        get_simul_beacon_data(beacon_data,index)
        simul_list = []
        for beacon in simul_beacon_data:
            for i in range(0,beaconNum):
                if beacon.uuid == test_beacons[i].uuid and \
                beacon.major == test_beacons[i].major and \
                beacon.minor == test_beacons[i].minor:
                    if i not in simul_list:
                        sim_mat[i][i] = sim_mat[i][i] + 1
                        simul_list.append(i)
        permu = two_permu(simul_list)
        for permu_item in permu:
            sim_mat[permu_item[0]][permu_item[1]] = \
            sim_mat[permu_item[0]][permu_item[1]] + 1
            
            
    # calculate the similarity score according to number of appearance
    for i in range(0,beaconNum):
        for j in range(i+1,beaconNum):
            sco_mat[i][j] = sim_mat[i][j]*2/(sim_mat[i][i]+sim_mat[j][j])
    
    return sim_mat,sco_mat

def writeToFile(fileName,trace_latitude,trace_longitude,trace_floor,\
                trace_timeStamp,trace_GPS_latitude,trace_GPS_longitude,\
                rider_id,date):
    import csv
    trace_Beacon = [["name","latitude","longitude","floor","timeStamp","type"]]
    trace_GPS = [["name","latitude","longitude","timeStamp","type"]]
    trace_sample = [["rider_id","date","latitude_beacon","longitude_beacon","floor",\
                     "timeStamp","latitude_GPS","longitude_GPS"]]
    for i in range(0,len(trace_latitude)):
        trace_Beacon.append([i+1,trace_latitude[i],trace_longitude[i],\
                      trace_floor[i],trace_timeStamp[i],"R"])
        trace_GPS.append([i+1,trace_GPS_latitude[i],\
                      trace_GPS_longitude[i],trace_timeStamp[i],"R"])
        trace_sample.append([rider_id,date,trace_latitude[i],trace_longitude[i],\
                      trace_floor[i],trace_timeStamp[i],trace_GPS_latitude[i],\
                      trace_GPS_longitude[i]])
    s1 = "."
    s2 = "_GPS."
    s3 = "_"
    with open(s1.join([fileName,"csv"]),'w') as f:
        writer = csv.writer(f)
        writer.writerows(trace_Beacon)
    with open(s2.join([fileName,"csv"]),'w') as f:
        writer = csv.writer(f)
        writer.writerows(trace_GPS)
    with open(s1.join([s3.join([fileName,"show"]),"csv"]),'w') as f:
        writer = csv.writer(f)
        writer.writerows(trace_sample)
        
        
def writeToArvFile(figureName,shop_lat,shop_lon,beacon_lat,beacon_lon,\
                   GPS_lat,GPS_lon,arr_order,date,arr_time,shop_id,rider_id,\
                   shop_floor,beacon_floor):
    import csv
    arrv_record = [["order_id","date","time","shop_id","rider_id","shop_latitude","shop_longitude",\
                    "beacon_latitude","beacon_longitude",\
                    "GPS_latitude","GPS_longitude","shop_floor","beacon_floor"]]
    for i in range(0,len(beacon_lat)):
        arrv_record.append([arr_order[i],date,arr_time[i],shop_id,rider_id,\
                            shop_lat,shop_lon,beacon_lat[i],beacon_lon[i],\
                            GPS_lat[i],GPS_lon[i],shop_floor,beacon_floor[i]])
    s1 = "."
    s3 = "_"
    with open(s1.join([s3.join([figureName,"show"]),"csv"]),'w')as f:
        writer = csv.writer(f)
        writer.writerows(arrv_record)
    
# This function draw the check in figure and write to the file
def draw_check_in(beacon_lat,beacon_lon,GPS_lat,GPS_lon,shop_lat,shop_lon,\
                  date,shop_id,rider_id,arr_order,arr_time,shop_floor,beacon_floor):
    import gmplot
    
    if len(beacon_lat) == 0:
        raise Exception("Empty beacon_lat?!")
    
    center_lat = beacon_lat[0]
    center_lon = beacon_lon[0]
    
    gmap = gmplot.GoogleMapPlotter(center_lat, center_lon, 16)
    gmap.scatter(beacon_lat, beacon_lon, 'blue', size=20, marker=True)
    gmap.scatter(GPS_lat, GPS_lon, 'red', size=20, marker=True)
    gmap.scatter([shop_lat], [shop_lon],'black', size=60, marker=True)
    figureName = '_'.join(['trace/arrival/arrival_comp',str(date),str(shop_id),str(rider_id)])
    figureFile = '.'.join([figureName,'html'])
    
    writeToArvFile(figureName,shop_lat,shop_lon,beacon_lat,beacon_lon,GPS_lat,\
                   GPS_lon,arr_order,date,arr_time,shop_id,rider_id,shop_floor,\
                   beacon_floor)
    gmap.draw(figureFile)

# This function draw the trace figure and write to the file
def draw_trace(center_lat, center_lon, whole_trace_lat,whole_trace_lon,\
               whole_trace_GPS_lat,whole_trace_GPS_lon,timeStamp,startTime,\
               timeBound,duration,rider_id,date):
    
    endTime = timeAdd(startTime, duration)
    index = 0
    
    endTime = timeAdd(startTime,duration)
    
    while index < len(timeStamp) and time_delta(timeBound,timeStamp[index]) > 0:
        # Vaild data can be fetched
        trace_lat = []
        trace_lon = []
        trace_GPS_lat = []
        trace_GPS_lon = []
        while index < len(timeStamp) and time_delta(timeStamp[index],endTime) < 0:
            # The entry corresponding to index is in the current time interval
            trace_lat.append(whole_trace_lat[index])
            trace_lon.append(whole_trace_lon[index])
            trace_GPS_lat.append(whole_trace_GPS_lat[index])
            trace_GPS_lon.append(whole_trace_GPS_lon[index])
            index = index + 1            
        
        if len(trace_lat) == 0 :
            #print('No trace in current interval. From',startTime ,'to', endTime)
            startTime = endTime
            endTime = timeAdd(endTime, duration)
            continue
        #print("StartTime: ",startTime)
        
        gmap = gmplot.GoogleMapPlotter(center_lat, center_lon, 16)
        gmap.plot(trace_lat, trace_lon, 'blue', edge_width=10)
        gmap.scatter(trace_lat, trace_lon, 'blue', marker=True)
        # gmap.plot(trace_latitude_3, trace_longitude_3, 'yellow', edge_width=10)
        # gmap.scatter(trace_latitude_3, trace_longitude_3, 'yellow', marker=True)
        gmap.plot(trace_GPS_lat, trace_GPS_lon, 'red', edge_width=10)
        gmap.scatter(trace_GPS_lat, trace_GPS_lon, 'red', marker=True)
        # gmap.plot(trace_latitude_truth, trace_longitude_truth, 'maroon', edge_width=10)
        # gmap.scatter(trace_latitude_truth, trace_longitude_truth, 'maroon', marker=True)
        hs1 = str(startTime[0])
        hm = str(startTime[1])
        hs2 = str(startTime[2])
        comma = ':'
        time = comma.join([hs1,hm,hs2])
        sep = '_'
        fileName = sep.join(["trace/trace_comp",str(rider_id),str(date),time,".html"])      
        gmap.draw(fileName)
        #print("plot done")
        startTime = endTime
        endTime = timeAdd(endTime, duration)
        
#        print("StartTime after: ",startTime)
#        print("endTime after: ",endTime)
#        print("index after: ",index)
#        print("timeBound after: ",timeBound)
#        print("timeStamp[index]: ",timeStamp[index])
    
    
def trace_time_align(timeStamp,raw_timeStamp,raw_trace_lat,raw_trace_lon,\
                     raw_trace_floor):
    if len(raw_timeStamp) < len(timeStamp):
        raise Exception("raw data is shorter?!")
    
    raw_lat_shrk = [None for i in range(0,len(timeStamp))]
    raw_lon_shrk = [None for i in range(0,len(timeStamp))]
    raw_flr_shrk = [None for i in range(0,len(timeStamp))]
    j = 1
    for i in range(0,len(timeStamp)):
        
        has_find = 0
        while has_find == 0:
            if j == (len(raw_timeStamp)-1):
                raw_lat_shrk[i] = raw_trace_lat[j]
                raw_lon_shrk[i] = raw_trace_lon[j]
                raw_flr_shrk[i] = raw_trace_floor[j]
                has_find = 1
                break
            
            tm_dlt_pre = abs(time_delta(timeStamp[i],raw_timeStamp[j-1]))
            tm_dlt_crt = abs(time_delta(timeStamp[i],raw_timeStamp[j]))
            tm_dlt_nxt = abs(time_delta(timeStamp[i],raw_timeStamp[j+1]))
            
            # print("at i: ",i,"tm_dlt_pre: ",tm_dlt_pre,"tm_dlt_crt",tm_dlt_crt,"tm_dlt_nxt",tm_dlt_nxt)
            
            if tm_dlt_crt < tm_dlt_pre and tm_dlt_crt < tm_dlt_nxt:
                raw_lat_shrk[i] = raw_trace_lat[j]
                raw_lon_shrk[i] = raw_trace_lon[j]
                raw_flr_shrk[i] = raw_trace_floor[j]
                has_find = 1
            elif tm_dlt_pre < tm_dlt_crt and tm_dlt_pre < tm_dlt_nxt:
                raw_lat_shrk[i] = raw_trace_lat[j-1]
                raw_lon_shrk[i] = raw_trace_lon[j-1]
                raw_flr_shrk[i] = raw_trace_floor[j-1]
                has_find = 1
                
            j = j + 1

    if len(timeStamp) != len(raw_lat_shrk):
        print("len(timeStamp): ",len(timeStamp))
        print("len(traw_lat_shrk): ",len(raw_lat_shrk))
        raise Exception("Not same size after shrink?!")
    return raw_lat_shrk,raw_lon_shrk,raw_flr_shrk

# Generic trace draw function
def generic_trace_draw(cen_lat,cen_lon,num_trace,trace_lat,trace_lon,trace_flr,\
                       trace_time,start_time,bound_time,duration,rider_id,date):
    if num_trace != len(trace_lat):
        raise Exception("Wrong trace num!")
    color = ["red","blue","yellow","maroon","aqua"]
        
    index = [0 for i in range(0,num_trace)]

    # End time of each duration    
    end_time = timeAdd(start_time, duration)

    while time_delta(bound_time,end_time) > 0:
        # While there are still some plots to draw
        lat_duration = [[] for i in range(0, num_trace)]
        lon_duration = [[] for i in range(0, num_trace)]
        flr_duration = [[] for i in range(0, num_trace)]
        
        gmap = gmplot.GoogleMapPlotter(cen_lat, cen_lon, 16)
        
        no_data = 1
        
        for trace_i in range(0,num_trace):
            while index[trace_i] < len(trace_time[trace_i]) \
            and time_delta(trace_time[trace_i][index[trace_i]],end_time) < 0:
                # The index[trace_i] is valid and corresponding time is within 
                # the interval
                lat_duration[trace_i].append(trace_lat[trace_i][index[trace_i]])
                lon_duration[trace_i].append(trace_lon[trace_i][index[trace_i]])
                flr_duration[trace_i].append(trace_flr[trace_i][index[trace_i]])                
                
                index[trace_i] = index[trace_i] + 1
                
            if len(lat_duration[0]) == 0:
                #print('No data for trace ',trace_i, ' from ', \
                #      start_time, ' to ', end_time)
                continue
            
            no_data = 0
            
            gmap.plot(lat_duration[trace_i], lon_duration[trace_i], color[trace_i], edge_width=10)            
            gmap.scatter(lat_duration[trace_i], lon_duration[trace_i], color[trace_i], marker=True)

            
        if no_data == 1:
            start_time = end_time
            end_time = timeAdd(end_time, duration)
            continue
            
        hs1 = str(start_time[0])
        hm = str(start_time[1])
        hs2 = str(start_time[2])
        comma = ':'
        time = comma.join([hs1,hm,hs2])
        sep = '_'
        fileName = sep.join(["trace/trace_comp",str(rider_id),str(date),time,".html"])      
        gmap.draw(fileName)
        print("plot done")
        print("index: ",index)
            
        start_time = end_time
        end_time = timeAdd(end_time, duration)
            
            
# This function draw the trace figure and write to the file
def draw_trace_three(center_lat, center_lon, whole_trace_lat,whole_trace_lon,\
               whole_trace_GPS_lat,whole_trace_GPS_lon,whole_raw_trace_lat,\
               whole_raw_trace_lon,timeStamp,startTime,\
               timeBound,duration,rider_id,date):
    
    endTime = timeAdd(startTime, duration)
    index = 0
    
    
    while index < len(timeStamp) and time_delta(timeBound,timeStamp[index]) > 0:
        # Vaild data can be fetched
        trace_lat = []
        trace_lon = []
        trace_GPS_lat = []
        trace_GPS_lon = []
        trace_raw_lat = []
        trace_raw_lon = []
        while index < len(timeStamp) and time_delta(timeStamp[index],endTime) < 0:
            # The entry corresponding to index is in the current time interval
            trace_lat.append(whole_trace_lat[index])
            trace_lon.append(whole_trace_lon[index])
            trace_GPS_lat.append(whole_trace_GPS_lat[index])
            trace_GPS_lon.append(whole_trace_GPS_lon[index])
            trace_raw_lat.append(whole_raw_trace_lat[index])
            trace_raw_lon.append(whole_raw_trace_lon[index])
            index = index + 1            
        
        if len(trace_lat) == 0 :
            print('No trace in current interval. From',startTime ,'to', endTime)
            startTime = endTime
            endTime = timeAdd(endTime, duration)
            continue
        print("StartTime: ",startTime)
        
        gmap = gmplot.GoogleMapPlotter(center_lat, center_lon, 16)
        gmap.plot(trace_lat, trace_lon, 'blue', edge_width=10)
        gmap.scatter(trace_lat, trace_lon, 'blue', marker=True)
        gmap.plot(trace_raw_lat, trace_raw_lon, 'yellow', edge_width=10)
        gmap.scatter(trace_raw_lat, trace_raw_lon, 'yellow', marker=True)
        gmap.plot(trace_GPS_lat, trace_GPS_lon, 'red', edge_width=10)
        gmap.scatter(trace_GPS_lat, trace_GPS_lon, 'red', marker=True)
        # gmap.plot(trace_latitude_truth, trace_longitude_truth, 'maroon', edge_width=10)
        # gmap.scatter(trace_latitude_truth, trace_longitude_truth, 'maroon', marker=True)
        hs1 = str(startTime[0])
        hm = str(startTime[1])
        hs2 = str(startTime[2])
        comma = ':'
        time = comma.join([hs1,hm,hs2])
        sep = '_'
        fileName = sep.join(["trace/trace_comp",str(rider_id),str(date),time,".html"])      
        gmap.draw(fileName)
        print("plot done")
        startTime = endTime
        endTime = timeAdd(endTime, duration)
        
#        print("StartTime after: ",startTime)
#        print("endTime after: ",endTime)
#        print("index after: ",index)
#        print("timeBound after: ",timeBound)
#        print("timeStamp[index]: ",timeStamp[index])
        
# if two beacon that are more than 200 meters are heared within 5 seconds
# the two beacon will be marked as "exception"
def get_exceptional_shops(delta,distance,beacon_data,beacon_info):

    step_check = 5  # Check 5 steps  

    exception_tuple = []
    num_occur = []
    exception_minor = []
    for i in range(0,len(beacon_data)):
        for j in range(1,step_check):
            if i+j == len(beacon_data):
                break
            bd1 = beacon_data[i]
            bd2 = beacon_data[i+j]
            if time_delta([bd2.lclHour,bd2.lclMnut,bd2.lclScnd],\
                         [bd1.lclHour,bd1.lclMnut,bd1.lclScnd]) > delta:
                continue
            if time_delta([bd2.lclHour,bd2.lclMnut,bd2.lclScnd],\
                         [bd1.lclHour,bd1.lclMnut,bd1.lclScnd]) < 0:
                print("i,j is: ",i,j)
                raise Exception("beacon_data Not time order?")
            
            if bd1.uuid == "E2C56DB5-DFFB-48D2-B060-D0F5A71096E0" and \
            bd2.uuid == "E2C56DB5-DFFB-48D2-B060-D0F5A71096E0":
                b1 = get_beacon_by_major_minor(bd1.major,bd1.minor,beacon_info)
                b2 = get_beacon_by_major_minor(bd2.major,bd2.minor,beacon_info)
                if b1 == None or b2 == None:
                    continue
                if b1.latitude == None or b2.latitude == None:
                    continue
                dist = calDist([b1.latitude,b1.longitude],[b2.latitude,b2.longitude])*100000
                if dist > distance:
                    shop1 = b1.shop_id
                    shop2 = b2.shop_id
                    if shop1 == None or shop2 == None:
                        print("bd1.major, bd2.major, bd1.minor, bd2.minor",\
                              bd1.major,bd2.major,bd1.minor,bd2.minor)
                        print("b1.major, b2.major, b1.minor, b2.minor",\
                              b1.major,b2.major,b1.minor,b2.minor)
                        raise Exception("No shop ID?")
                    minor1 = bd1.minor
                    minor2 = bd2.minor
                    if shop1 > shop2:
                        shop1 = b2.shop_id
                        shop2 = b1.shop_id
                        minor1 = bd2.minor
                        minor2 = bd1.minor
                    if [shop1,shop2] in exception_tuple:
                        num_occur[exception_tuple.index([shop1,shop2])] = \
                        num_occur[exception_tuple.index([shop1,shop2])] + 1
                    else:
                        exception_tuple.append([shop1,shop2])
                        num_occur.append(1) 
                        exception_minor.append([minor1,minor2])
                        
    return exception_tuple, exception_minor