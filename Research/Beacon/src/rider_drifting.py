"""
Python codes for rider drifting related issues
For Phase III reporting
Yi DING
02/26/18

"""

import os
import openpyxl
import gmplot

# How did we find the shop id?

# shop_id = 155307294
shop_id = 1778360

# This path of current file
dir_path = os.path.dirname(os.path.realpath(__file__))
data_path = os.path.join(dir_path, '../data/')
data_path = '/Users/eleme-yi/Downloads/beacon-temp'
figure_path = '/Users/eleme-yi/Downloads/beacon-temp'

# Set file name
file_name = '_'.join(['rider_drifting_shop', str(shop_id)])
file_name = '.'.join([file_name, 'xlsx'])
file_path = os.path.join(data_path, file_name)

# Open workbook
wb = openpyxl.load_workbook(file_path)
sheet = wb.get_sheet_by_name('Sheet0')
row_count = sheet.max_row
column_count = sheet.max_column

# Get argument list
argument_list = []
for j in range(1, column_count+1):
    argument_list.append(sheet.cell(row=1,column=j).value)

print('argument list: ', argument_list)

gps_lat_idx = argument_list.index('gps_latitude')
gps_lon_idx = argument_list.index('gps_longitude')
bcn_lat_idx = argument_list.index('beacon_latitude')
bcn_lon_idx = argument_list.index('beacon_longitude')

gps_lat_list = []
gps_lon_list = []
bcn_lat_list = []
bcn_lon_list = []

for i in range(2, row_count):
    gps_lat_list.append(sheet.cell(row=i, column=gps_lat_idx + 1).value)
    gps_lon_list.append(sheet.cell(row=i, column=gps_lon_idx + 1).value)
    bcn_lat_list.append(sheet.cell(row=i, column=bcn_lat_idx + 1).value)
    bcn_lon_list.append(sheet.cell(row=i, column=bcn_lon_idx + 1).value)

# Draw
gmap = gmplot.GoogleMapPlotter(bcn_lat_list[0], bcn_lon_list[0], 16)
gmap.scatter(gps_lat_list, gps_lon_list, 'blue', marker=True, size = 40)
gmap.scatter(bcn_lat_list, bcn_lon_list, 'red', marker=True, size = 80)
file_name = '_'.join(["rider_drifting_shop_",str(shop_id), ".html"])
file_path = os.path.join(figure_path, file_name)
gmap.draw(file_path)
print("plot done")
