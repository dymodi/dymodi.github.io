#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 27 14:15:40 2017

Shanghai shop clustering

@author: eleme-yi
"""

import os
import openpyxl
import gmplot
import numpy as np
from sklearn.preprocessing import scale
from sklearn.cluster import KMeans

# Some standard path operations
dir_path = os.path.dirname(os.path.realpath(__file__))
dir_root = os.path.dirname(dir_path)
dir_figure = os.path.join(dir_root, 'figures/')
dir_data = os.path.join(dir_root, 'data/shop/')

sep = '_'

file_name = 'indoor_shop_gps.xlsx'
file_path = os.path.join(dir_data,file_name)


wb = openpyxl.load_workbook(file_path)
sheet = wb.get_sheet_by_name('Sheet0')
row_count = sheet.max_row

shop_lat_list = []
shop_lon_list = []
shop_pos_list = []

# Read lat/lon from file
for i in range(2, row_count):
    lat_str = sheet.cell(row=i,column=3).value
    lon_str = sheet.cell(row=i,column=4).value
    shop_lat_list.append(float(lat_str))
    shop_lon_list.append(float(lon_str))
    shop_pos_list.append([float(lat_str),float(lon_str)])

# Data preprocessing    
shop_pos_list = np.array(shop_pos_list)
data = scale(shop_pos_list)

# clustering using kmeans
n_groups = 10
kmeans = KMeans(init='random', n_clusters=n_groups, n_init=10)
kmeans.fit(data)
labels = kmeans.labels_

colors = ['blue','red','green','cyan','black','maroon','olive','navy',\
          'purple','yellow']

gmap = gmplot.GoogleMapPlotter(31.2325, 121.381895, 16)

for i in range(0, len(data)):
    gmap.scatter([shop_lat_list[i]], [shop_lon_list[i]], colors[labels[i]], \
                 marker=True)

output_file = os.path.join(dir_figure,'indoor_shop_cluster.html')
gmap.draw(output_file)