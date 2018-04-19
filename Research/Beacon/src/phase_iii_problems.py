"""
This script is used to discover and solve problems for Ele.Me beacon Phase III
Yi Ding
02/20/2018

"""

import os
import openpyxl
from datetime import datetime
from pytz import timezone
import gmplot

## Draw shops whose beacon are heard at the same second
# This path of current file
dir_path = os.path.dirname(os.path.realpath(__file__))
data_path = os.path.join(dir_path, '../data/')
figure_path = os.path.join(dir_path, '../figures/')

# # Read raw beacon data for specific (rider, shop) pair
# rider_id = 10646314
# date_str = '2018-02-05'
# time_query = '2018-02-05 10:37:00'
# Below is a recent iOS rider
rider_id = 102134170
date_str = '2018-02-23'
time_query = '2018-02-23 13:56:31'


# Set file name
file_name = '_'.join(['dw_ai_clairvoyant_beacon', date_str, str(rider_id)])
file_name = '.'.join([file_name, 'xlsx'])
file_path = os.path.join(data_path, file_name)

# Open workbook
wb = openpyxl.load_workbook(file_path)
sheet = wb.get_sheet_by_name('Sheet0')
row_count = sheet.max_row
column_count = sheet.max_column

# Get argument list
argument_list = []
for j in range(1, column_count+1):
    argument_list.append(sheet.cell(row=1,column=j).value)

detected_at_index = argument_list.index('detected_at')
shop_index = argument_list.index('shop_id')
lat_index = argument_list.index('latitude')
lon_index = argument_list.index('longitude')

# Time conversion
datetime_query = datetime.strptime(time_query, "%Y-%m-%d %H:%M:%S")
datetime_query_SH = timezone('Asia/Shanghai').localize(datetime_query)
timestamp_query =int(datetime_query_SH.timestamp())

# Get lists from raw file
lat_list = []
lon_list = []
shop_list = []
for i in range(2, row_count+1):
    detected_time = str(sheet.cell(row=i, column=detected_at_index+1).value)
    datetime_detected = datetime.strptime(detected_time, "%Y-%m-%d %H:%M:%S")
    datetime_detected_SH = timezone('Asia/Shanghai').localize(datetime_detected)
    timestamp_detected = int(datetime_detected_SH.timestamp())
    if timestamp_query == timestamp_detected:
        shop_list.append(sheet.cell(row=i, column=shop_index+1).value)
        lat_list.append(sheet.cell(row=i, column=lat_index+1).value)
        lon_list.append(sheet.cell(row=i, column=lon_index+1).value)

# Plot lat/lon
gmap = gmplot.GoogleMapPlotter(lat_list[0], lon_list[0], 16)
gmap.scatter(lat_list, lon_list, 'blue', marker=True)
file_name = '_'.join(["beacons_at_same_time",str(rider_id),time_query,".html"])
file_path = os.path.join(figure_path, file_name)
gmap.draw(file_path)
print("plot done")