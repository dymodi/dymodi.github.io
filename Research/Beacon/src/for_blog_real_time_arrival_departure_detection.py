# This is for the blog "Real-Time Arrival Departure Detection"
# 02/04/18 Yi DING

# Import
import openpyxl
import os
from datetime import datetime
from pytz import timezone
import matplotlib.pyplot as plt
from scipy import signal
import numpy as np
from radd import window_filter as window_filter

# # Draw RSSI - Distance relation
# rssi_list = []
# dist_list = []
# A = -55
# n = 2
# for rssi in range(-100, -50):
#     rssi_list.append(rssi)
#     dist = np.power(10, (-rssi+A)/(10*n))
#     dist_list.append(dist)
#
# # Plot In region result after shave
# plt.plot(dist_list, rssi_list)
# plt.xlabel('distance / meters')
# plt.ylabel('RSSI / dB ')
# plt.title('RSSI - Distance relation')
# plt.show()


# This path of current file
dir_path = os.path.dirname(os.path.realpath(__file__))
data_path = os.path.join(dir_path, '../data/')

# Read raw beacon data for specific (rider, shop) pair
rider_id = 102081898
shop_id = 72887
date_str = '2018-02-18'
date_str_start = "2018-02-18 12:55:00"
date_str_end = '2018-02-18 13:05:00'
# Time conversion
time_start = datetime.strptime(date_str_start, "%Y-%m-%d %H:%M:%S")
time_end = datetime.strptime(date_str_end, "%Y-%m-%d %H:%M:%S")
time_start_SH = timezone('Asia/Shanghai').localize(time_start)
time_end_SH = timezone('Asia/Shanghai').localize(time_end)
timestamp_start =int(time_start_SH.timestamp())
timestamp_end =int(time_end_SH.timestamp())


# # # Possible samples:
# rider_id = 100556090
# shop_id = 147525716
# date_str = '2018-02-04'
# date_str_start = "2018-02-04 11:00:00"
# date_str_end = '2018-02-04 11:10:00'
# # # Possible samples:
# rider_id = 100556090
# shop_id = 159134974
# date_str = '2018-02-04'
# date_str_start = "2018-02-04 12:20:00"
# date_str_end = '2018-02-04 12:30:00'

# Read beacon data file
# Set file name
file_name = '_'.join(['dw_ai_clairvoyant_beacon', date_str, str(rider_id)])
file_name = '.'.join([file_name, 'xlsx'])
file_path = os.path.join(data_path, file_name)
# Open workbook
wb = openpyxl.load_workbook(file_path)
sheet = wb.get_sheet_by_name('Sheet0')
row_count = sheet.max_row
column_count = sheet.max_column
# Get argument list
argument_list = []
for j in range(1, column_count+1):
    argument_list.append(sheet.cell(row=1,column=j).value)
print('Beacon data file argument list: ', argument_list)
# Read by row
rssi_list = []
timestamp_list = []
shop_list = []
shop_id_list = []
rssi_index = argument_list.index('rssi')
shop_index = argument_list.index('shop_id')
timestamp_index = argument_list.index('unix_timestamp')
shop_index = argument_list.index('shop_id')
for i in range(2, row_count+1):
    unix_timestamp = int(sheet.cell(row=i, column=timestamp_index+1).value)
    if unix_timestamp < timestamp_start or unix_timestamp > timestamp_end:
        continue
    rssi_list.append(sheet.cell(row=i, column=rssi_index+1).value)
    timestamp_list.append(unix_timestamp)
    shop_id = int(sheet.cell(row=i, column=shop_index+1).value)
    shop_list.append(shop_id)
    if shop_id not in shop_id_list:
        shop_id_list.append(shop_id)
print(len(rssi_list), ' RSSI signals got.')

# Read rider's event file (to get ground truth)
# Set file name
file_name = '_'.join(['dw_tms_tb_tracking_event', date_str, str(rider_id)])
file_name = '.'.join([file_name, 'xlsx'])
file_path = os.path.join(data_path, file_name)
# Open workbook
wb = openpyxl.load_workbook(file_path)
sheet = wb.get_sheet_by_name('Sheet0')
row_count = sheet.max_row
column_count = sheet.max_column
# Get argument list
argument_list = []
for j in range(1, column_count+1):
    argument_list.append(sheet.cell(row=1,column=j).value)
print('Rider event argument list: ', argument_list)


# Pick RSSI from specific time slot
rssi_draw = []
timestamp_draw = []
for timestp in range(timestamp_start, timestamp_end):
    timestamp_draw.append(timestp)
    if timestp in timestamp_list and shop_id == shop_list[timestamp_list.index(timestp)]:
        rssi_draw.append(rssi_list[timestamp_list.index(timestp)])
    else:
        rssi_draw.append(np.nan)


# Low Pass Filter Way of RADD
# First Here we design a low-pass filter
fs = 1
nyq = 0.5 * fs
# wn = 0.05 / nyq
wn = 0.05
order = 8
b, a = signal.butter(order, wn, 'low', analog = False)

rssi_filtered = signal.filtfilt(b, a, rssi_draw)

# Arrival/Departure Detection with RSSI threshold
in_region_raw = []
rssi_thresh = -100
for rssi in rssi_filtered:
    if rssi > rssi_thresh:
        in_region_raw.append(1)
    else:
        in_region_raw.append(0)

# Shave
in_region = window_filter(in_region_raw, 20)

# Design a moving filter and work on RSSI slot
horizon_length = 30
i = 0
rssi_range_filtered = []
while i + horizon_length < len(rssi_draw):
    rssi_window = rssi_draw[i:i+horizon_length]
    rssi_window_filtered = signal.filtfilt(b, a, rssi_window)
    for rssi in rssi_window_filtered:
        rssi_range_filtered.append(rssi)
    i = i + horizon_length

# Result of short range LPF after shaving
# Plot Arrival/Departure Detection with RSSI threshold
in_region_online = []
for rssi in rssi_range_filtered:
    if rssi > rssi_thresh:
        in_region_online.append(1)
    else:
        in_region_online.append(0)
# Shave
in_region_online_shaved = window_filter(in_region_online, 20)

# # Plot Original RSSI in the time slot
# plt.plot(timestamp_draw, rssi_draw, 'C0')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('RSSI / dB (absent set as -100dB)')
# plt.title('Rider\' RSSI value in 10 minutes for specific shop ')
# plt.show()
#
# # Plot filtered RSSI
# plt.plot(timestamp_draw, rssi_filtered, 'C1')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('RSSI / dB (absent set as -100dB)')
# plt.title('Rider\' RSSI value in 10 minutes for specific shop (Offline recognition) ')
# plt.show()
#
# # Plot In region result
# plt.plot(timestamp_draw, in_region_raw, 'C2')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('In Region?')
# plt.title('Rider\'s in region information (Offline recognition)')
# plt.show()
#
# # Plot In region result after shave
# plt.plot(timestamp_draw, in_region, 'C3')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('In Region?')
# plt.title('Rider\'s in region information after shave (Offline recognition)')
# plt.show()
#
# # Plot RSSI with short range LPF
# len_filtered = len(rssi_range_filtered)
# plt.plot(timestamp_draw[0:len_filtered], rssi_range_filtered, 'C4')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('RSSI / dB (absent set as -100dB)')
# plt.title('Rider\' RSSI value in 10 minutes for specific shop (Real-time recognition)')
# plt.show()
#
# # Plot In region result
# plt.plot(timestamp_draw[0:len_filtered], in_region_online, 'C5')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('In Region?')
# plt.title('Rider\'s in region information (Real-time recognition)')
# plt.show()
#
# # Plot In region result after shave
# plt.plot(timestamp_draw[0:len_filtered], in_region_online_shaved, 'C6')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('In Region?')
# plt.title('Rider\'s in region information after shave (Real-time recognition')
# plt.show()

# -----------------------------------------------------------------------------
# Single plot done

# Pick RSSI from specific time slot
time_axis = range(timestamp_start, timestamp_end+1)
rssi_matrix = np.full((len(time_axis), len(shop_id_list)), np.nan)
for i in range(0,len(rssi_list)):
    # What shop is it?
    shop_id = shop_list[i]
    shop_index = shop_id_list.index(shop_id)
    # What time is it?
    timestamp = timestamp_list[i]
    time_index = time_axis.index(timestamp)
    # Assign RSSI to the position
    rssi_matrix[time_index][shop_index] = rssi_list[i]

# Plot rider's RSSI with respect to multiple shops
for i in range(0, len(shop_id_list)):
    #plt.plot(time_axis, rssi_matrix[:,i], label = 'shop'+str(i))
    plt.plot(time_axis, rssi_matrix[:, i], '-*', markersize=6, markeredgewidth=0, label='shop_id: '+str(shop_id_list[i]))
plt.xlabel('unix_timestamp / second')
plt.ylabel('RSSI / dB (absent set as -100dB)')
plt.title('Rider\'s in region information (Real-time recognition)')
plt.legend()
plt.show()

# Suppose we have RSSI from m shops across N time steps.
# size(time_axis) = [1, N], unix_timestamps
# size(rssi_matrix) = [N, m], RSSI values from m shops at each step


# -----------------------------------------------------------------------------
# Conduct Bayesian Estimation on Rider's Prescence
#radd.bayesian(time_axis, shop_id_list, rssi_matrix)


